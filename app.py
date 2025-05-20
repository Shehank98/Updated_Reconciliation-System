import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import re
import plotly.express as px
import plotly.graph_objects as go
from matplotlib import pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from datetime import datetime
import os

def create_summary_pdf_tables_only(summary_data, channel_name, matched_df, schedule_df, unmatched_schedule_df=None):
    """Create a PDF summary report with detailed tables for media reconciliation"""
    import os
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    # Create a unique filename
    filename = f"media_reconciliation_summary_{channel_name}{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"

    # Create PDF document
    doc = SimpleDocTemplate(filename, pagesize=letter)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading1"]
    subheading_style = styles["Heading2"]
    normal_style = styles["Normal"]

    # Add title and date
    elements.append(Paragraph(f"Media Reconciliation Summary - {channel_name}", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 12))

    # Add summary statistics
    elements.append(Paragraph("Summary Statistics", heading_style))

    if matched_df is not None and schedule_df is not None:
        match_rate = (len(matched_df)/max(len(schedule_df), 1))*100 if schedule_df is not None else 0
        stats_data = [
            ["Metric", "Value"],
            ["Total LMRB Records", str(len(matched_df))],
            ["Total Schedule Spots", str(len(schedule_df)) if schedule_df is not None else "0"],
            ["Match Rate", f"{match_rate:.2f}%"]
        ]

        # If we have theme mapping data, add theme count
        if 'Media_Watch_Theme' in matched_df.columns:
            theme_count = matched_df['Media_Watch_Theme'].nunique()
            stats_data.append(["Unique Themes", str(theme_count)])

        t = Table(stats_data, colWidths=[250, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

    elements.append(Spacer(1, 24))

    
    # 2. THEME DISTRIBUTION - Now grouped by duration 
    if matched_df is not None and 'Media_Watch_Theme' in matched_df.columns:
        elements.append(Paragraph("Theme Distribution Summary", heading_style))
        
        try:
            # Find duration column
            duration_col = None
            for col in ['Dur', 'Duration', 'DURATION']:
                if col in matched_df.columns:
                    duration_col = col
                    break
                    
            if duration_col:
                # Group by theme and duration
                theme_counts = matched_df.groupby(['Schedule_Theme', duration_col]).size().reset_index(name='Count')
                theme_counts = theme_counts.sort_values('Count', ascending=False).head(15)  # Top 15 theme-duration combinations
                
                # Add percentage column
                total_spots = theme_counts['Count'].sum()
                theme_counts['Percentage'] = (theme_counts['Count'] / matched_df.shape[0] * 100).round(1)
                
                theme_table_data = [["Theme Name", "Duration", "Count", "Percentage"]]
                for _, row in theme_counts.iterrows():
                    theme_table_data.append([
                        row['Schedule_Theme'], 
                        str(row[duration_col]), 
                        str(row['Count']), 
                        f"{row['Percentage']}%"
                    ])
                    
                t = Table(theme_table_data, colWidths=[200, 50, 100, 100])
            else:
                # Without duration, just group by theme
                theme_counts = matched_df['Media_Watch_Theme'].value_counts().reset_index()
                theme_counts.columns = ['Theme', 'Count']
                theme_counts['Percentage'] = (theme_counts['Count'] / matched_df.shape[0] * 100).round(1)
                theme_counts = theme_counts.head(10)
                
                theme_table_data = [["Theme Name", "Count", "Percentage"]]
                for _, row in theme_counts.iterrows():
                    theme_table_data.append([row['Theme'], str(row['Count']), f"{row['Percentage']}%"])
                    
                t = Table(theme_table_data, colWidths=[250, 100, 100])
                
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),  # Right align count & percentage columns
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t)
        except Exception as e:
            elements.append(Paragraph(f"Error creating theme distribution: {str(e)}", normal_style))
            
        elements.append(Spacer(1, 12))



    # SCHEDULE COMPLIANCE SUMMARY - Fixed with dynamic column detection
    if matched_df is not None and schedule_df is not None and 'Schedule_Theme' in matched_df.columns:
        elements.append(PageBreak())  # Start on a new page
        elements.append(Paragraph("Schedule Compliance Summary", heading_style))
        
        try:
            # Find duration column in each dataset
            matched_dur_col = None
            schedule_dur_col = None
            
            for col in ['Dur', 'Duration', 'DURATION']:
                if col in matched_df.columns:
                    matched_dur_col = col
                if col in schedule_df.columns:
                    schedule_dur_col = col
            
            if matched_dur_col and schedule_dur_col:
                # Create copies to avoid modifying originals
                matched_copy = matched_df.copy()
                schedule_copy = schedule_df.copy()
                
                # Group by theme and duration
                grouped = matched_copy.groupby(['Schedule_Theme', matched_dur_col]).size().reset_index(name='Aired_Count')
                schedule_counts = schedule_copy.groupby(['Advt_Theme', schedule_dur_col]).size().reset_index(name='Scheduled_Count')
                
                # Rename columns for consistent merging
                schedule_counts.rename(columns={'Advt_Theme': 'Theme', schedule_dur_col: 'Dur'}, inplace=True)
                grouped.rename(columns={'Schedule_Theme': 'Theme', matched_dur_col: 'Dur'}, inplace=True)
                
                # Merge datasets
                merged = pd.merge(schedule_counts, grouped, on=['Theme', 'Dur'], how='left')
            else:
                # Fall back to theme-only grouping
                grouped = matched_df.groupby(['Schedule_Theme']).size().reset_index(name='Aired_Count')
                schedule_counts = schedule_df.groupby(['Advt_Theme']).size().reset_index(name='Scheduled_Count')
                schedule_counts.rename(columns={'Advt_Theme': 'Theme'}, inplace=True)
                grouped.rename(columns={'Schedule_Theme': 'Theme'}, inplace=True)
                merged = pd.merge(schedule_counts, grouped, on='Theme', how='left')
            
            merged['Aired_Count'].fillna(0, inplace=True)
            merged['Aired_Count'] = merged['Aired_Count'].astype(int)
            merged['Missed_Count'] = merged['Scheduled_Count'] - merged['Aired_Count']
            merged['Missed_Count'] = merged['Missed_Count'].clip(lower=0)
            merged['Compliance_Rate'] = (merged['Aired_Count'] / merged['Scheduled_Count'] * 100).round(1)
            
            # Calculate duration metrics if duration column exists
            if 'Dur' in merged.columns:
                merged['Planned_Duration_Secs'] = merged['Scheduled_Count'] * merged['Dur']
                merged['Planned_30s_Equiv'] = (merged['Planned_Duration_Secs'] / 30).round(1)
            
            # Sort by compliance rate and get top themes
            top_themes = merged.sort_values('Compliance_Rate').head(20)

            # Create appropriate table structure - Use only requested columns
            if 'Dur' in merged.columns:
                compliance_table_data = [["Theme", "Duration", "Scheduled", "Aired", "Missed", "30s Equiv", "Compliance %"]]
                for _, row in top_themes.iterrows():
                    compliance_table_data.append([
                        row['Theme'],
                        str(row['Dur']),
                        str(row['Scheduled_Count']),
                        str(row['Aired_Count']),
                        str(row['Missed_Count']),
                        str(row['Planned_30s_Equiv']),
                        f"{row['Compliance_Rate']}%"
                    ])
                
                t = Table(compliance_table_data, colWidths=[150, 50, 60, 60, 60, 60, 80])
            else:
                compliance_table_data = [["Theme", "Scheduled", "Aired", "Missed", "Compliance %"]]
                for _, row in top_themes.iterrows():
                    compliance_table_data.append([
                        row['Theme'],
                        str(row['Scheduled_Count']),
                        str(row['Aired_Count']),
                        str(row['Missed_Count']),
                        f"{row['Compliance_Rate']}%"
                    ])
                
                t = Table(compliance_table_data, colWidths=[200, 70, 70, 70, 90])

            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

        except Exception as e:
            elements.append(Paragraph(f"Error creating compliance table: {str(e)}", normal_style))

# AD POSITION ANALYSIS - FOCUSED ON POSITIONS 1 AND 2 ONLY
# Let's check for AdPos column in the dataset
        if matched_df is not None and any(col in matched_df.columns for col in ['AdPos', 'Ad_Position', 'Position']):
            elements.append(Paragraph("Ad Position Analysis", heading_style))

        try:
            pos_df = matched_df.copy()
            
            # Find the position column
            pos_column = None
            for col in ['AdPos', 'Ad_Position', 'Position']:
                if col in pos_df.columns:
                    pos_column = col
                    break
            
            if pos_column:
                # Convert position to numeric if needed
                try:
                    pos_df[pos_column] = pd.to_numeric(pos_df[pos_column], errors='coerce')
                except:
                    pass
                    
                # Count total spots
                total_spots = len(pos_df)
                
                # Debug: Print position counts
                position_counts = pos_df[pos_column].value_counts().to_dict()
                position1_count = position_counts.get(1, 0)
                position2_count = position_counts.get(2, 0)
                other_positions_count = total_spots - position1_count - position2_count
                
                # Calculate percentages
                position1_pct = round((position1_count / total_spots) * 100, 1) if total_spots > 0 else 0
                position2_pct = round((position2_count / total_spots) * 100, 1) if total_spots > 0 else 0
                other_pct = round((other_positions_count / total_spots) * 100, 1) if total_spots > 0 else 0
                
                # Create focused table data
                pos_table_data = [
                    ["Ad Position", "Count", "Percentage"],
                    ["Position 1", str(position1_count), f"{position1_pct}%"],
                    ["Position 2", str(position2_count), f"{position2_pct}%"],
                    ["Other Positions", str(other_positions_count), f"{other_pct}%"],
                    ["Total", str(total_spots), "100.0%"]
                ]
                
                # Create table
                t = Table(pos_table_data, colWidths=[150, 100, 100])
                
                # Style the table with highlighting for positions 1 and 2
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    # Highlight rows for positions 1 and 2
                    ('BACKGROUND', (0, 1), (-1, 1), colors.lightcoral),
                    ('BACKGROUND', (0, 2), (-1, 2), colors.lightcoral)
                ]
                
                t.setStyle(TableStyle(table_style))
                elements.append(t)
                
                # Add explanatory note about positions 1 and 2
                elements.append(Spacer(1, 6))
                premium_positions_pct = position1_pct + position2_pct
                
                highlight_note = (
                    f"Note: Premium positions (1 and 2) account for {premium_positions_pct:.1f}% of all ad positions. "
                    f"Position 1 represents {position1_pct:.1f}% and Position 2 represents {position2_pct:.1f}% of total spots."
                )
                
                elements.append(Paragraph(highlight_note, normal_style))
                
                # Add additional insight if applicable
                if premium_positions_pct > 50:
                    elements.append(Paragraph(
                        "Insight: The majority of ads appear in premium positions, indicating favorable ad placement.",
                        normal_style
                    ))
                elif premium_positions_pct < 30:
                    elements.append(Paragraph(
                        "Insight: Less than a third of ads appear in premium positions, suggesting opportunity for improved positioning.",
                        normal_style
                    ))
                
            else:
                elements.append(Paragraph("Ad position column not found in data", normal_style))
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            elements.append(Paragraph(f"Error creating ad position analysis: {str(e)}", normal_style))
            elements.append(Paragraph(f"Error details: {error_details}", normal_style))

        elements.append(Spacer(1, 12))

    # 8. MISSED SPOTS DETAILS - On a separate page
    if unmatched_schedule_df is not None and not unmatched_schedule_df.empty:
        elements.append(PageBreak())
        elements.append(Paragraph("Missed Spots Details", heading_style))
        
        try:
            # Show missed spots with just the requested columns
            missed_spots = unmatched_schedule_df.head(15)  # Show more missed spots
            
            # Only include specific columns
            cols_to_show = ['Advt_Theme', 'Program', 'Date', 'Advt_time']
            cols_to_show = [c for c in cols_to_show if c in missed_spots.columns]
            
            # Format date if needed
            if 'Date' not in missed_spots.columns and all(col in missed_spots.columns for col in ['Dd', 'Mn', 'Yr']):
                missed_spots = missed_spots.copy()
                missed_spots['Date'] = missed_spots.apply(
                    lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                    axis=1
                )
                cols_to_show.append('Date')
            
            missed_table = [cols_to_show]
            for _, row in missed_spots.iterrows():
                missed_table.append([str(row[col]) for col in cols_to_show])
            
            # Create table with appropriate column widths
            col_widths = []
            for col in cols_to_show:
                if col == 'Advt_Theme':
                    col_widths.append(100)
                elif col == 'Program':
                    col_widths.append(100)
                else:
                    col_widths.append(100)
            
            t = Table(missed_table, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            elements.append(t)
            
            # Add summary metrics for missed spots
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("Missed Spots Summary", subheading_style))
            
            # Group missed spots by theme for an additional table
            if 'Advt_Theme' in unmatched_schedule_df.columns:
                theme_missed = unmatched_schedule_df['Advt_Theme'].value_counts().reset_index()
                theme_missed.columns = ['Theme', 'Missed_Count']
                theme_missed = theme_missed.sort_values('Missed_Count', ascending=False).head(10)
                
                # Create table
                theme_missed_table = [["Theme", "Missed Count"]]
                for _, row in theme_missed.iterrows():
                    theme_missed_table.append([row['Theme'], str(row['Missed_Count'])])
                
                theme_missed_t = Table(theme_missed_table, colWidths=[300, 100])
                theme_missed_t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(theme_missed_t)
            
            # Add missed spots summary paragraph
            elements.append(Spacer(1, 8))
            missed_summary = f"Total of {len(unmatched_schedule_df)} spots were missed in the schedule. "
            
            # Group by program to add information
            if 'Program' in unmatched_schedule_df.columns:
                program_counts = unmatched_schedule_df['Program'].value_counts().head(3)
                missed_summary += "Most missed spots occurred in programs: "
                missed_summary += ", ".join([f"{program} ({count} spots)" for program, count in program_counts.items()])
                
            elements.append(Paragraph(missed_summary, normal_style))
            
        except Exception as e:
            elements.append(Paragraph(f"Error creating missed spots table: {str(e)}", normal_style))

    # Add notes section
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Notes", heading_style))
    elements.append(Paragraph(
        "This report provides a comprehensive analysis of media reconciliation including day of week distribution, "
        "prime time analysis, ad position analysis, and program performance. "
        "For interactive charts and full visual analysis, please refer to the Streamlit application dashboard.", 
        normal_style
    ))

    # Build PDF
    doc.build(elements)
    return filename

def display_chart(chart, title=None, container_width=True, key_prefix="chart"):
    """Display a Plotly chart with a unique key to avoid duplicate element IDs"""
    import random
    import string

    # Generate random suffix for key uniqueness
    suffix = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
    unique_key = f"{key_prefix}_{suffix}"

    if title:
        st.subheader(title)

    # Correctly call st.plotly_chart directly (not display_chart recursively)
    st.plotly_chart(chart, use_container_width=container_width, key=unique_key)

def extract_duration_from_theme(theme_name):
    """Extract duration in seconds from theme names like 'Cash Bonanza_2025 (10)(Sin)'"""
    if pd.isna(theme_name):
        return None

    # Look for patterns like (10), (15), (30) in theme names
    duration_match = re.search(r'(\d+)(?=\))', str(theme_name))
    if duration_match:
        return int(duration_match.group(1))
    return None

def preprocess_time_format(time_value):
    """Extract and standardize time format from various inputs."""
    try:
        if isinstance(time_value, str):
            time_pattern = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)', str(time_value))
            if time_pattern:
                time_str = time_pattern.group(1)
                if time_str.count(':') == 1:
                    time_str += ':00'
                return time_str
            else:
                return '00:00:00'
        elif hasattr(time_value, 'strftime'):
            return time_value.strftime('%H:%M:%S')
        return '00:00:00'
    except Exception:
        return '00:00:00'

def time_to_seconds(time_str):
    """Convert time string (HH:MM:SS) to seconds."""
    try:
        if pd.isna(time_str):
            return 0
            
        time_str = str(time_str)
        time_parts = time_str.split(':')
        
        if len(time_parts) >= 3:
            return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60 + int(time_parts[2])
        elif len(time_parts) == 2:
            return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60
        else:
            return 0
        
    except (ValueError, IndexError):
        return 0

def normalize_theme_name(theme):
    """Normalize theme name for consistent comparison by removing whitespace and lowercase."""
    if pd.isna(theme):
        return ""
    return str(theme).lower().strip()

def filter_by_channel(df, channel):
    """Filter data by selected channel."""
    if "Channel" not in df.columns:
        st.error("Channel column not found in the dataframe.")
        return df
    return df[df["Channel"] == channel].reset_index(drop=True)

def remove_duplicates(current_df, previous_df):
    """Remove duplicates between current and previous month's data."""
    current_df = current_df.reset_index(drop=True)
    previous_df = previous_df.reset_index(drop=True)
    match_columns = ["Advt_Theme", "Channel", "Program", "Advt_time", "Dd", "Mn", "Yr"]

    for col in match_columns:
        if col not in current_df.columns or col not in previous_df.columns:
            st.warning(f"Column '{col}' missing in one of the dataframes. Skipping duplicate removal.")
            return current_df, 0
        
    merged_df = pd.merge(
        current_df, previous_df, 
        on=match_columns,
        how="left", indicator=True
    )
    cleaned_df = merged_df[merged_df["_merge"] == "left_only"].drop(columns=["_merge"]).reset_index(drop=True)
    duplicate_count = len(current_df) - len(cleaned_df)

    return cleaned_df, duplicate_count

def clean_column_names(df):
    """Clean up column names by removing _x and _y suffixes from merged dataframes."""
    col_mapping = {}
    seen_base_cols = set()

    for col in df.columns:
        if col.endswith('_x') or col.endswith('_y'):
            base_name = col[:-2]
            if base_name not in seen_base_cols:
                col_mapping[col] = base_name
                seen_base_cols.add(base_name)
        else:
            col_mapping[col] = col

    if 'Advt_time' in df.columns and df.columns.tolist().count('Advt_time') > 1:
        advt_time_indices = [i for i, col in enumerate(df.columns) if col == 'Advt_time']
        for i, idx in enumerate(advt_time_indices[1:], 1):
            original_name = df.columns[idx]
            new_name = f'Advt_time_{i}'
            col_mapping[original_name] = new_name

    return df.rename(columns=col_mapping)

def standardize_dataframe(df, date_col=None, time_col=None, theme_col=None, program_col=None):
    """Standardize DataFrame columns and formats for consistent processing."""
    df = clean_column_names(df)
    result_df = df.copy().reset_index(drop=True)

    rename_dict = {}
    if theme_col and theme_col in df.columns:
        rename_dict[theme_col] = 'Advt_Theme'
    if program_col and program_col in df.columns:
        rename_dict[program_col] = 'Program'
    if time_col and time_col in df.columns:
        rename_dict[time_col] = 'Advt_time'

    if rename_dict:
        result_df = result_df.rename(columns=rename_dict)

    if date_col and date_col in df.columns:
        result_df['Date'] = pd.to_datetime(df[date_col], errors='coerce')
        result_df['Dd'] = result_df['Date'].dt.day
        result_df['Mn'] = result_df['Date'].dt.month
        result_df['Yr'] = result_df['Date'].dt.year

    if 'Advt_time' in result_df.columns:
        result_df['Advt_time'] = result_df['Advt_time'].apply(preprocess_time_format)
        
    if 'Advt_Theme' in result_df.columns:
        result_df['Normalized_Theme'] = result_df['Advt_Theme'].apply(normalize_theme_name)
        # Extract duration from theme names
        result_df['Theme_Duration'] = result_df['Advt_Theme'].apply(extract_duration_from_theme)

    return result_df

def match_media_watch_with_tc(media_watch_df, tc_df, theme_mapping, ignore_date=False, time_tolerance=30):
    """Match Media Watch data with TC data based on theme mapping."""
    def local_time_to_seconds(time_str):
        try:
            if pd.isna(time_str):
                return 0
            time_str = str(time_str)
            time_parts = time_str.split(':')
            if len(time_parts) >= 3:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60 + int(time_parts[2])
            elif len(time_parts) == 2:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60
            else:
                return 0
        except (ValueError, IndexError):
            return 0

    if media_watch_df.empty or tc_df.empty:
        st.error("One or both DataFrames are empty.")
        return pd.DataFrame(), [], pd.DataFrame()

    # Find and fix Advt_time in TC data
    time_related_cols = [col for col in tc_df.columns if any(term in str(col).lower() for term in ['time', 'spot', 'air'])]

    # Check for Advt_time with exact capitalization
    if 'Advt_time' not in tc_df.columns:
        # Try to find case-insensitive match
        for col in tc_df.columns:
            if col.lower() == 'advt_time':
                tc_df = tc_df.rename(columns={col: 'Advt_time'})
                break
        else:
            # If no exact match found, use the first time-related column
            if time_related_cols:
                tc_df = tc_df.rename(columns={time_related_cols[0]: 'Advt_time'})

    media_watch_df = media_watch_df.copy().reset_index(drop=True)
    tc_df = tc_df.copy().reset_index(drop=True)

    if 'Advt_time' in media_watch_df.columns:
        media_watch_df['Time_Seconds'] = media_watch_df['Advt_time'].astype(str).apply(local_time_to_seconds)
    else:
        st.error("Advt_time column not found in Media Watch data")
        return pd.DataFrame(), [], pd.DataFrame()
        
    if 'Advt_time' in tc_df.columns:
        tc_df['Time_Seconds'] = tc_df['Advt_time'].astype(str).apply(local_time_to_seconds)
    else:
        st.error("Advt_time column not found in TC data")
        return pd.DataFrame(), [], pd.DataFrame()

    mw_to_tc = {}
    for m in theme_mapping:
        if 'tc_theme' in m and m['tc_theme']:
            mw_theme = normalize_theme_name(m['media_watch_theme'])
            tc_theme = normalize_theme_name(m['tc_theme'])
            mw_to_tc[mw_theme] = tc_theme

    progress_bar = st.progress(0)
    total_rows = len(media_watch_df)

    matched_results = []
    matched_mw_indices = []
    time_matched_program_mismatched = []

    filter_stats = {
        "total_media_watch_rows": len(media_watch_df),
        "no_normalized_mapping": 0,
        "no_matches_after_theme_filter": 0,
        "no_matches_after_date_filter": 0,
        "no_matches_after_time_filter": 0,
        "matches_found": 0,
        "failed_program_similarity": 0,
        "program_mismatched_but_time_matched": 0
    }

    for i in range(len(media_watch_df)):
        progress_bar.progress(min(1.0, (i + 1) / total_rows))
        mw_row = media_watch_df.iloc[i]
        
        if pd.isna(mw_row['Advt_Theme']) or pd.isna(mw_row['Program']) or pd.isna(mw_row['Advt_time']):
            continue
        
        mw_theme_norm = normalize_theme_name(mw_row['Advt_Theme'])
        
        if mw_theme_norm not in mw_to_tc:
            filter_stats["no_normalized_mapping"] += 1
            continue
            
        tc_theme_norm = mw_to_tc[mw_theme_norm]
        tc_matches = tc_df[tc_df['Normalized_Theme'] == tc_theme_norm].copy().reset_index(drop=True)
        
        if tc_matches.empty:
            filter_stats["no_matches_after_theme_filter"] += 1
            continue
            
        # Extract duration from theme
        mw_duration = extract_duration_from_theme(mw_row['Advt_Theme'])
        
        # Filter by duration if available
        if mw_duration is not None:
            if 'Theme_Duration' in tc_matches.columns:
                # First try theme duration from extracted theme name
                duration_matches = tc_matches[tc_matches['Theme_Duration'] == mw_duration]
                if not duration_matches.empty:
                    tc_matches = duration_matches.reset_index(drop=True)
            elif 'Dur' in tc_matches.columns:
                # Then try explicit duration column
                duration_matches = tc_matches[tc_matches['Dur'] == mw_duration]
                if not duration_matches.empty:
                    tc_matches = duration_matches.reset_index(drop=True)
        
        if not ignore_date:
            date_filter = (
                (tc_matches['Dd'] == mw_row['Dd']) & 
                (tc_matches['Mn'] == mw_row['Mn']) & 
                (tc_matches['Yr'] == mw_row['Yr'])
            )
            tc_matches = tc_matches[date_filter].reset_index(drop=True)
        
        if tc_matches.empty:
            filter_stats["no_matches_after_date_filter"] += 1
            continue
            
        mw_time_seconds = mw_row['Time_Seconds']
        tc_matches['Time_Diff'] = abs(tc_matches['Time_Seconds'] - mw_time_seconds)
        
        # Increased time tolerance to 30 seconds (configurable)
        close_time_matches = tc_matches[tc_matches['Time_Diff'] <= time_tolerance].reset_index(drop=True)
        
        if close_time_matches.empty:
            filter_stats["no_matches_after_time_filter"] += 1
            continue
            
        close_time_matches = close_time_matches.sort_values('Time_Diff').reset_index(drop=True)
        close_time_matches['Program_Similarity'] = close_time_matches['Program'].astype(str).apply(
            lambda x: fuzz.token_set_ratio(str(mw_row['Program']).lower(), x.lower())
        )
        
        # Handle multiple potential matches
        if len(close_time_matches) > 1:
            # Sort by program similarity first, then time difference
            close_time_matches = close_time_matches.sort_values(
                ['Program_Similarity', 'Time_Diff'], 
                ascending=[False, True]
            ).reset_index(drop=True)
        
        if close_time_matches.empty:
            filter_stats["no_matches_after_time_filter"] += 1
            continue
            
        best_match = close_time_matches.iloc[0]
        
        # Check if time matches but program doesn't match well
        if best_match['Program_Similarity'] < 50:
            # This is a time-matched but program-mismatched case
            filter_stats["program_mismatched_but_time_matched"] += 1
            
            # Create a result for program mismatched but time matched cases
            mismatch_row = {
                'Media_Watch_Theme': mw_row['Advt_Theme'],
                'TC_Theme': best_match['Advt_Theme'],
                'Date': f"{mw_row['Dd']}/{mw_row['Mn']}/{mw_row['Yr']}",
                'Dd': mw_row['Dd'],
                'Mn': mw_row['Mn'],
                'Yr': mw_row['Yr'],
                'Program_LMRB': mw_row['Program'],
                'Program_TC': best_match['Program'],
                'Media_Watch_Time': mw_row['Advt_time'],
                'TC_Time': best_match['Advt_time'],
                'Program_Similarity': best_match['Program_Similarity'],
                'Time_Difference_Seconds': best_match['Time_Diff'],
                'Channel': mw_row.get('Channel', ''),
                'Match_Status': 'Time Matched / Program Mismatched'
            }
            time_matched_program_mismatched.append(mismatch_row)
            continue
                
        result_row = {
            'Media_Watch_Theme': mw_row['Advt_Theme'],
            'TC_Theme': best_match['Advt_Theme'],
            'Date': f"{mw_row['Dd']}/{mw_row['Mn']}/{mw_row['Yr']}",
            'Dd': mw_row['Dd'],
            'Mn': mw_row['Mn'],
            'Yr': mw_row['Yr'],
            'Program': mw_row['Program'],
            'TC_Program': best_match['Program'],
            'Media_Watch_Time': mw_row['Advt_time'],
            'TC_Time': best_match['Advt_time'],
            'Program_Similarity': best_match['Program_Similarity'],
            'Time_Difference_Seconds': best_match['Time_Diff'],
            'Channel': mw_row.get('Channel', ''),
            'Advt_time': mw_row['Advt_time'],
            'Match_Status': 'Full Match',
            'Dur': mw_row.get('Dur', mw_duration)  # Add duration for schedule matching
        }
        
        # Copy all original LMRB columns
        for col in media_watch_df.columns:
            if col not in result_row and col not in ['Normalized_Theme', 'Time_Seconds', 'Theme_Duration']:
                col_name = f'LMRB_{col}' if col in result_row else col
                result_row[col_name] = mw_row[col]
        
        # Copy all TC columns
        for col in tc_df.columns:
            if col not in result_row and col not in ['Normalized_Theme', 'Time_Seconds', 'Time_Diff', 'Program_Similarity', 'Theme_Duration']:
                col_name = f'TC_{col}' if col in result_row else col
                result_row[col_name] = best_match[col]
                
        filter_stats["matches_found"] += 1
        matched_results.append(result_row)
        matched_mw_indices.append(i)

    progress_bar.empty()

    matched_df = pd.DataFrame(matched_results) if matched_results else pd.DataFrame()
    program_mismatched_df = pd.DataFrame(time_matched_program_mismatched) if time_matched_program_mismatched else pd.DataFrame()

    return matched_df, matched_mw_indices, program_mismatched_df

def match_media_watch_with_schedule(media_watch_df, schedule_df, theme_mapping, ignore_date=False):
    """Match Media Watch data directly with Schedule data when TC theme is empty."""
    def local_time_to_seconds(time_str):
        try:
            if pd.isna(time_str):
                return 0
            time_str = str(time_str)
            time_parts = time_str.split(':')
            if len(time_parts) >= 3:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60 + int(time_parts[2])
            elif len(time_parts) == 2:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60
            else:
                return 0
        except (ValueError, IndexError):
            return 0

    if media_watch_df.empty or schedule_df.empty:
        st.error("One or both DataFrames are empty.")
        return pd.DataFrame(), []

    media_watch_df = media_watch_df.copy().reset_index(drop=True)
    schedule_df = schedule_df.copy().reset_index(drop=True)

    if 'Advt_time' in media_watch_df.columns:
        media_watch_df['Time_Seconds'] = media_watch_df['Advt_time'].astype(str).apply(local_time_to_seconds)
        
    if 'Advt_time' in schedule_df.columns:
        schedule_df['Time_Seconds'] = schedule_df['Advt_time'].astype(str).apply(local_time_to_seconds)

    mw_to_schedule = {}
    for m in theme_mapping:
        if 'schedule_theme' in m and m['schedule_theme'] and ('tc_theme' not in m or not m['tc_theme']):
            mw_theme = normalize_theme_name(m['media_watch_theme'])
            schedule_theme = normalize_theme_name(m['schedule_theme'])
            mw_to_schedule[mw_theme] = schedule_theme

    progress_bar = st.progress(0)
    total_rows = len(media_watch_df)

    matched_results = []
    matched_mw_indices = []

    filter_stats = {
        "total_media_watch_rows": len(media_watch_df),
        "no_normalized_mapping": 0,
        "no_matches_after_theme_filter": 0,
        "no_matches_after_date_filter": 0,
        "no_matches_after_duration_filter": 0,
        "matches_found": 0,
        "failed_program_similarity": 0
    }

    for i in range(len(media_watch_df)):
        progress_bar.progress(min(1.0, (i + 1) / total_rows))
        mw_row = media_watch_df.iloc[i]
        
        if pd.isna(mw_row['Advt_Theme']) or pd.isna(mw_row['Program']):
            continue
        
        mw_theme_norm = normalize_theme_name(mw_row['Advt_Theme'])
        
        if mw_theme_norm not in mw_to_schedule:
            filter_stats["no_normalized_mapping"] += 1
            continue
            
        schedule_theme_norm = mw_to_schedule[mw_theme_norm]
        schedule_matches = schedule_df[schedule_df['Normalized_Theme'] == schedule_theme_norm].copy().reset_index(drop=True)
        
        if schedule_matches.empty:
            filter_stats["no_matches_after_theme_filter"] += 1
            continue
            
        # Extract duration from theme
        mw_duration = extract_duration_from_theme(mw_row['Advt_Theme'])
        
        # Filter by duration if available
        if mw_duration is not None:
            if 'Theme_Duration' in schedule_matches.columns:
                # First try theme duration from extracted theme name
                duration_matches = schedule_matches[schedule_matches['Theme_Duration'] == mw_duration]
                if not duration_matches.empty:
                    schedule_matches = duration_matches.reset_index(drop=True)
            elif 'Dur' in schedule_matches.columns:
                # Then try explicit duration column
                duration_matches = schedule_matches[schedule_matches['Dur'] == mw_duration]
                if not duration_matches.empty:
                    schedule_matches = duration_matches.reset_index(drop=True)
        
        if not ignore_date:
            date_filter = (
                (schedule_matches['Dd'] == mw_row['Dd']) & 
                (schedule_matches['Mn'] == mw_row['Mn']) & 
                (schedule_matches['Yr'] == mw_row['Yr'])
            )
            schedule_matches = schedule_matches[date_filter].reset_index(drop=True)
        
        if schedule_matches.empty:
            filter_stats["no_matches_after_date_filter"] += 1
            continue
            
        if 'Dur' in mw_row and 'Dur' in schedule_matches.columns:
            mw_duration = mw_row['Dur']
            schedule_matches['Duration_Diff'] = abs(schedule_matches['Dur'] - mw_duration)
            close_dur_matches = schedule_matches[schedule_matches['Duration_Diff'] <= 1].reset_index(drop=True)
            
            if close_dur_matches.empty:
                filter_stats["no_matches_after_duration_filter"] += 1
                continue
                
            schedule_matches = close_dur_matches.sort_values('Duration_Diff').reset_index(drop=True)
        
        schedule_matches['Program_Similarity'] = schedule_matches['Program'].astype(str).apply(
            lambda x: fuzz.token_set_ratio(str(mw_row['Program']).lower(), x.lower())
        )
        
        if 'Duration_Diff' in schedule_matches.columns:
            schedule_matches = schedule_matches.sort_values(
                ['Program_Similarity', 'Duration_Diff'], 
                ascending=[False, True]
            ).reset_index(drop=True)
        else:
            schedule_matches = schedule_matches.sort_values(
                'Program_Similarity', 
                ascending=False
            ).reset_index(drop=True)
        
        if schedule_matches.empty:
            continue
            
        best_match = schedule_matches.iloc[0]
        
        if best_match['Program_Similarity'] < 50:
            filter_stats["failed_program_similarity"] += 1
            continue
            
        result_row = {
            'Media_Watch_Theme': mw_row['Advt_Theme'],
            'TC_Theme': '',
            'Schedule_Theme': best_match['Advt_Theme'],
            'Date': f"{mw_row['Dd']}/{mw_row['Mn']}/{mw_row['Yr']}",
            'Dd': mw_row['Dd'],
            'Mn': mw_row['Mn'],
            'Yr': mw_row['Yr'],
            'Program': mw_row['Program'],
            'Schedule_Program': best_match['Program'],
            'Media_Watch_Time': mw_row.get('Advt_time', ''),
            'Program_Similarity': best_match['Program_Similarity'],
            'Channel': mw_row.get('Channel', ''),
            'Match_Status': 'Direct Schedule Match'
        }
        
        if 'Duration_Diff' in best_match:
            result_row['Duration_Difference'] = best_match['Duration_Diff']
        
        # Add original LMRB columns
        for col in media_watch_df.columns:
            if col not in result_row and col not in ['Normalized_Theme', 'Time_Seconds', 'Theme_Duration']:
                col_name = f'LMRB_{col}' if col in result_row else col
                result_row[col_name] = mw_row[col]
        
        # Add schedule columns
        for col in schedule_df.columns:
            if col not in result_row and col not in ['Normalized_Theme', 'Duration_Diff', 'Program_Similarity', 'Time_Seconds', 'Theme_Duration']:
                col_name = f'Schedule_{col}' if col in result_row else col
                result_row[col_name] = best_match[col]
        
        filter_stats["matches_found"] += 1
        matched_results.append(result_row)
        matched_mw_indices.append(i)

    progress_bar.empty()

    matched_df = pd.DataFrame(matched_results) if matched_results else pd.DataFrame()
    return matched_df, matched_mw_indices

def match_with_schedule(matched_mw_tc_df, schedule_df, theme_mapping, ignore_date=False):
    """Match the MW+TC matched data with Schedule data based on theme mapping."""
    def local_time_to_seconds(time_str):
        try:
            if pd.isna(time_str):
                return 0
            time_str = str(time_str)
            time_parts = time_str.split(':')
            if len(time_parts) >= 3:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60 + int(time_parts[2])
            elif len(time_parts) == 2:
                return int(time_parts[0]) * 3600 + int(time_parts[1]) * 60
            else:
                return 0
        except (ValueError, IndexError):
            return 0

    if matched_mw_tc_df.empty or schedule_df.empty:
        st.error("One or both DataFrames are empty.")
        return pd.DataFrame(), pd.DataFrame()

    matched_mw_tc_df = matched_mw_tc_df.copy().reset_index(drop=True)
    schedule_df = schedule_df.copy().reset_index(drop=True)

    if 'Advt_time' in matched_mw_tc_df.columns and 'Time_Seconds' not in matched_mw_tc_df.columns:
        matched_mw_tc_df['Time_Seconds'] = matched_mw_tc_df['Advt_time'].astype(str).apply(local_time_to_seconds)
        
    if 'Advt_time' in schedule_df.columns and 'Time_Seconds' not in schedule_df.columns:
        schedule_df['Time_Seconds'] = schedule_df['Advt_time'].astype(str).apply(local_time_to_seconds)

    mw_to_schedule = {}
    tc_to_schedule = {}

    for m in theme_mapping:
        if 'schedule_theme' in m and m['schedule_theme']:
            if 'media_watch_theme' in m and m['media_watch_theme']:
                mw_theme = normalize_theme_name(m['media_watch_theme'])
                schedule_theme = normalize_theme_name(m['schedule_theme'])
                mw_to_schedule[mw_theme] = schedule_theme
            
            if 'tc_theme' in m and m['tc_theme']:
                tc_theme = normalize_theme_name(m['tc_theme'])
                schedule_theme = normalize_theme_name(m['schedule_theme'])
                tc_to_schedule[tc_theme] = schedule_theme

    progress_bar = st.progress(0)
    total_rows = len(matched_mw_tc_df)

    matched_with_schedule = []
    unmatched_with_schedule = []

    for i in range(len(matched_mw_tc_df)):
        progress_bar.progress(min(1.0, (i + 1) / total_rows))
        mw_tc_row = matched_mw_tc_df.iloc[i]
        
        tc_theme_norm = normalize_theme_name(mw_tc_row['TC_Theme'])
        schedule_theme_norm = None
        
        if tc_theme_norm in tc_to_schedule:
            schedule_theme_norm = tc_to_schedule[tc_theme_norm]
        elif 'Media_Watch_Theme' in mw_tc_row:
            mw_theme_norm = normalize_theme_name(mw_tc_row['Media_Watch_Theme'])
            if mw_theme_norm in mw_to_schedule:
                schedule_theme_norm = mw_to_schedule[mw_theme_norm]
        
        if not schedule_theme_norm:
            # No mapping found - add to unmatched
            row_copy = dict(mw_tc_row)
            row_copy['Match_Status'] = 'No Schedule Theme Mapping'
            unmatched_with_schedule.append(row_copy)
            continue
            
        schedule_matches = schedule_df[schedule_df['Normalized_Theme'] == schedule_theme_norm].copy().reset_index(drop=True)
        
        if schedule_matches.empty:
            row_copy = dict(mw_tc_row)
            row_copy['Match_Status'] = 'No Matching Schedule Theme'
            unmatched_with_schedule.append(row_copy)
            continue
            
        # Extract duration from theme
        if 'TC_Theme' in mw_tc_row:
            tc_duration = extract_duration_from_theme(mw_tc_row['TC_Theme'])
        
        # Filter by duration if available
        if tc_duration is not None:
            if 'Theme_Duration' in schedule_matches.columns:
                # First try theme duration from extracted theme name
                duration_matches = schedule_matches[schedule_matches['Theme_Duration'] == tc_duration]
                if not duration_matches.empty:
                    schedule_matches = duration_matches.reset_index(drop=True)
            elif 'Dur' in schedule_matches.columns:
                # Then try explicit duration column
                duration_matches = schedule_matches[schedule_matches['Dur'] == tc_duration]
                if not duration_matches.empty:
                    schedule_matches = duration_matches.reset_index(drop=True)
        
        if not ignore_date:
            date_filter = (
                (schedule_matches['Dd'] == mw_tc_row['Dd']) & 
                (schedule_matches['Mn'] == mw_tc_row['Mn']) & 
                (schedule_matches['Yr'] == mw_tc_row['Yr'])
            )
            schedule_matches = schedule_matches[date_filter].reset_index(drop=True)
        
        if schedule_matches.empty:
            row_copy = dict(mw_tc_row)
            row_copy['Match_Status'] = 'No Schedule Match on Date'
            row_copy['Different_Date'] = True  # Mark as aired on different date
            unmatched_with_schedule.append(row_copy)
            continue
            
        if 'Dur' in mw_tc_row and 'Dur' in schedule_matches.columns:
            mw_duration = mw_tc_row['Dur']
            schedule_matches['Duration_Diff'] = abs(schedule_matches['Dur'] - mw_duration)
            close_dur_matches = schedule_matches[schedule_matches['Duration_Diff'] <= 1].reset_index(drop=True)
            
            if not close_dur_matches.empty:
                schedule_matches = close_dur_matches
            else:
                row_copy = dict(mw_tc_row)
                row_copy['Match_Status'] = 'Duration Mismatch with Schedule'
                unmatched_with_schedule.append(row_copy)
                continue
        
        schedule_matches['Program_Similarity'] = schedule_matches['Program'].astype(str).apply(
            lambda x: fuzz.token_set_ratio(str(mw_tc_row['Program']).lower(), x.lower())
        )
        
        if 'Duration_Diff' in schedule_matches.columns:
            schedule_matches = schedule_matches.sort_values(
                ['Program_Similarity', 'Duration_Diff'], 
                ascending=[False, True]
            ).reset_index(drop=True)
        else:
            schedule_matches = schedule_matches.sort_values(
                'Program_Similarity', 
                ascending=False
            ).reset_index(drop=True)
        
        if schedule_matches.empty:
            row_copy = dict(mw_tc_row)
            row_copy['Match_Status'] = 'No Matching Schedule After Filtering'
            unmatched_with_schedule.append(row_copy)
            continue
            
        best_match = schedule_matches.iloc[0]
        
        # Flag for different program
        different_program = best_match['Program_Similarity'] < 70
        
        # Even if program similarity is low, still match but note the difference
        result_row = dict(mw_tc_row)
        result_row['Schedule_Theme'] = best_match['Advt_Theme']
        result_row['Schedule_Program'] = best_match['Program']
        result_row['Program_Similarity_Schedule'] = best_match['Program_Similarity']
        result_row['Different_Program'] = different_program
        
        if different_program:
            result_row['Match_Status'] = 'Match with Different Program'
        else:
            result_row['Match_Status'] = 'Full Match with Schedule'
            
        if 'Duration_Diff' in best_match:
            result_row['Duration_Difference'] = best_match['Duration_Diff']
        
        for col in schedule_df.columns:
            if col not in result_row and col not in ['Normalized_Theme', 'Duration_Diff', 'Program_Similarity', 'Time_Seconds', 'Theme_Duration']:
                col_name = f'Schedule_{col}'
                result_row[col_name] = best_match[col]
        
        matched_with_schedule.append(result_row)

    progress_bar.empty()
    matched_df = pd.DataFrame(matched_with_schedule) if matched_with_schedule else pd.DataFrame()
    unmatched_df = pd.DataFrame(unmatched_with_schedule) if unmatched_with_schedule else pd.DataFrame()

    return matched_df, unmatched_df

def generate_comprehensive_summary_report(matched_df, original_media_watch, original_schedule):
    """Generate a detailed reconciliation summary report with durations and calculations."""
    if matched_df.empty:
        return None

    # FIND DURATION COLUMN NAMES IN ALL DATASETS
    duration_col_matched = None
    duration_col_schedule = None

    # Check for duration column in matched data
    for col_name in ['Dur', 'Duration', 'DURATION']:
        if col_name in matched_df.columns:
            duration_col_matched = col_name
            break

    # Check for duration column in schedule data
    if original_schedule is not None:
        for col_name in ['Dur', 'Duration', 'DURATION']:
            if col_name in original_schedule.columns:
                duration_col_schedule = col_name
                break

    # First, make sure we have Schedule_Theme for proper grouping
    if 'Schedule_Theme' not in matched_df.columns:
        if 'Media_Watch_Theme' in matched_df.columns:
            matched_df['Schedule_Theme'] = matched_df['Media_Watch_Theme']

    # Group by Schedule theme and duration if available
    group_by_columns = []
    if 'Schedule_Theme' in matched_df.columns:
        group_by_columns.append('Schedule_Theme')

    if duration_col_matched:
        group_by_columns.append(duration_col_matched)
        # Standardize to 'Dur' for output
        if duration_col_matched != 'Dur':
            matched_df['Dur'] = matched_df[duration_col_matched]
            group_by_columns.remove(duration_col_matched)
            group_by_columns.append('Dur')

    # If no groupable columns, use Media_Watch_Theme
    if not group_by_columns and 'Media_Watch_Theme' in matched_df.columns:
        group_by_columns = ['Media_Watch_Theme']
        if 'Theme_Duration' in matched_df.columns:
            matched_df['Dur'] = matched_df['Theme_Duration']
            group_by_columns.append('Dur')

    # Basic summary with aired count
    if group_by_columns:
        summary = matched_df.groupby(group_by_columns).size().reset_index(name='Aired_Count')
    else:
        # Fallback if no groupable columns
        summary = pd.DataFrame({
            'Total': ['All Records'],
            'Aired_Count': [len(matched_df)]
        })

    # If schedule data is available, get scheduled counts
    if original_schedule is not None and 'Schedule_Theme' in summary.columns:
        if duration_col_schedule:
            # Standardize schedule duration column
            schedule_copy = original_schedule.copy()
            if duration_col_schedule != 'Dur':
                schedule_copy['Dur'] = schedule_copy[duration_col_schedule]
            
            # Group schedule by theme and duration
            schedule_summary = schedule_copy.groupby(['Advt_Theme', 'Dur']).size().reset_index(name='Planned_Count')
            schedule_summary.rename(columns={'Advt_Theme': 'Schedule_Theme'}, inplace=True)
            
            # Merge with summary
            summary = pd.merge(summary, schedule_summary, on=['Schedule_Theme', 'Dur'], how='left')
        else:
            # Group by theme only if no duration
            schedule_summary = original_schedule.groupby('Advt_Theme').size().reset_index(name='Planned_Count')
            schedule_summary.rename(columns={'Advt_Theme': 'Schedule_Theme'}, inplace=True)
            
            # Merge with summary
            summary = pd.merge(summary, schedule_summary, on='Schedule_Theme', how='left')

    # Fill NaN values and calculate metrics
    if 'Planned_Count' in summary.columns:
        summary['Planned_Count'] = summary['Planned_Count'].fillna(0).astype(int)
        summary['Missed_Count'] = summary['Planned_Count'] - summary['Aired_Count']
        summary['Missed_Count'] = summary['Missed_Count'].apply(lambda x: max(0, x))
        
        # Add extra spots count
        summary['Extra_Count'] = summary['Aired_Count'] - summary['Planned_Count']
        summary['Extra_Count'] = summary['Extra_Count'].apply(lambda x: max(0, x))
        
        # Calculate compliance percentage
        summary['Compliance_Rate'] = (summary['Aired_Count'] / summary['Planned_Count'] * 100).fillna(0)
        summary['Compliance_Rate'] = summary['Compliance_Rate'].clip(upper=100)

    # Calculate 30-second equivalent durations
    if 'Dur' in summary.columns:
        if 'Planned_Count' in summary.columns:
            summary['Planned_Duration_Secs'] = summary['Planned_Count'] * summary['Dur']
            summary['Planned_30s_Equiv'] = (summary['Planned_Duration_Secs'] / 30).round(1)
            
        summary['Aired_Duration_Secs'] = summary['Aired_Count'] * summary['Dur']
        summary['Aired_30s_Equiv'] = (summary['Aired_Duration_Secs'] / 30).round(1)
        
        if 'Missed_Count' in summary.columns:
            summary['Missed_Duration_Secs'] = summary['Missed_Count'] * summary['Dur']
            summary['Missed_30s_Equiv'] = (summary['Missed_Duration_Secs'] / 30).round(1)
        
        if 'Extra_Count' in summary.columns:
            summary['Extra_Duration_Secs'] = summary['Extra_Count'] * summary['Dur']
            summary['Extra_30s_Equiv'] = (summary['Extra_Duration_Secs'] / 30).round(1)

    # Add match status breakdown
    if 'Match_Status' in matched_df.columns:
        try:
            status_counts = matched_df.groupby(group_by_columns + ['Match_Status']).size().reset_index(name='Status_Count')
            
            # Create pivot to spread status counts across columns
            status_pivot = status_counts.pivot_table(
                index=group_by_columns,
                columns='Match_Status',
                values='Status_Count',
                fill_value=0
            ).reset_index()
            
            # Merge with summary
            if set(group_by_columns).issubset(set(summary.columns)):
                summary = pd.merge(summary, status_pivot, on=group_by_columns, how='left')
            
            # Fill NAs with 0
            for col in status_pivot.columns:
                if col not in group_by_columns and col in summary.columns:
                    summary[col] = summary[col].fillna(0).astype(int)
        except Exception:
            pass  # Skip status breakdown if pivot fails

    # Calculate totals for numeric columns
    numeric_cols = summary.select_dtypes(include=['int', 'float']).columns.tolist()
    summary_totals = summary[numeric_cols].sum().to_frame().T

    # Add placeholder values for non-numeric columns
    for col in summary.columns:
        if col not in numeric_cols:
            summary_totals[col] = "Total"

    # Reorder totals row columns to match summary
    summary_totals = summary_totals[summary.columns]

    # Append totals row
    summary = pd.concat([summary, summary_totals], ignore_index=True)

    return summary

def find_unmatched_schedule_spots(schedule_df, matched_schedule_themes):
    """Find spots in the schedule that weren't matched with LMRB or TC data."""
    if schedule_df is None or schedule_df.empty:
        return pd.DataFrame()

    # Create a copy of the schedule
    schedule_copy = schedule_df.copy()

    # If no matches, all schedule spots are unmatched
    if not matched_schedule_themes:
        schedule_copy['Match_Status'] = 'Not Found in TC or LMRB'
        return schedule_copy

    # Check which schedule spots aren't in the matched data
    schedule_copy['Is_Matched'] = False

    for theme, dates in matched_schedule_themes.items():
        for date_tuple in dates:
            mask = (
                (schedule_copy['Normalized_Theme'] == theme) &
                (schedule_copy['Dd'] == date_tuple[0]) &
                (schedule_copy['Mn'] == date_tuple[1]) &
                (schedule_copy['Yr'] == date_tuple[2])
            )
            schedule_copy.loc[mask, 'Is_Matched'] = True

    # Filter to only unmatched spots
    unmatched_spots = schedule_copy[~schedule_copy['Is_Matched']].copy()
    unmatched_spots['Match_Status'] = 'Missing Spot - Not Found in TC or LMRB'

    return unmatched_spots.drop('Is_Matched', axis=1)

def create_tc_vs_lmrb_excel_report(matched_df, program_mismatched_df, unmatched_df, time_matched_program_mismatched_df, media_watch_filtered, filename="tc_vs_lmrb_report.xlsx"):
    """Create a comprehensive Excel report with multiple sheets for TC vs LMRB matching."""
    wb = Workbook()

    # Create Summary Sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    # Add title and header
    summary_sheet.merge_cells('A1:H1')
    title_cell = summary_sheet['A1']
    title_cell.value = "TC vs LMRB Reconciliation Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Add metadata
    summary_sheet['A3'] = "Report Date:"
    summary_sheet['B3'] = pd.Timestamp.now().strftime("%Y-%m-%d")

    # Add match statistics
    summary_sheet['A5'] = "Match Statistics"
    summary_sheet['A5'].font = Font(bold=True)

    summary_sheet['A6'] = "Total LMRB Records:"
    summary_sheet['B6'] = len(media_watch_filtered) if media_watch_filtered is not None else 0

    summary_sheet['A7'] = "Full Matches Found:"
    summary_sheet['B7'] = len(matched_df) if matched_df is not None else 0

    summary_sheet['A8'] = "Program Mismatches:"
    summary_sheet['B8'] = len(program_mismatched_df) if program_mismatched_df is not None else 0

    summary_sheet['A9'] = "Time Matched / Program Mismatched:"
    summary_sheet['B9'] = len(time_matched_program_mismatched_df) if time_matched_program_mismatched_df is not None else 0

    summary_sheet['A10'] = "Unmatched TC Entries:"
    summary_sheet['B10'] = len(unmatched_df) if unmatched_df is not None else 0

    # Calculate match rate
    if media_watch_filtered is not None and len(media_watch_filtered) > 0:
        match_rate = (len(matched_df) / len(media_watch_filtered)) * 100 if matched_df is not None else 0
        summary_sheet['A11'] = "Match Rate:"
        summary_sheet['B11'] = f"{match_rate:.2f}%"

    # Add matched records summary by theme
    if matched_df is not None and not matched_df.empty:
        theme_counts = matched_df.groupby('Media_Watch_Theme').size().reset_index(name='Count')
        
        summary_sheet['A13'] = "Matches by Theme"
        summary_sheet['A13'].font = Font(bold=True)
        
        summary_sheet['A14'] = "Theme"
        summary_sheet['B14'] = "Count"
        summary_sheet['A14'].font = Font(bold=True)
        summary_sheet['B14'].font = Font(bold=True)
        
        for i, (theme, count) in enumerate(zip(theme_counts['Media_Watch_Theme'], theme_counts['Count']), 15):
            summary_sheet[f'A{i}'] = theme
            summary_sheet[f'B{i}'] = count

    # Format the summary sheet
    for col in ['A', 'B']:
        for row in range(1, 20):
            cell = summary_sheet[f'{col}{row}']
            cell.alignment = Alignment(vertical='center')
            if row in [5, 13, 14]:
                cell.font = Font(bold=True)

    # Adjust column widths - FIXED VERSION
    for col_idx in range(1, summary_sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for row_idx in range(1, summary_sheet.max_row + 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            if hasattr(cell, 'value') and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = max_length + 2
        summary_sheet.column_dimensions[column_letter].width = adjusted_width

    # Create Full Matches Sheet
    if matched_df is not None and not matched_df.empty:
        full_matches_sheet = wb.create_sheet("Full Matches")
        
        # Add headers
        for col_idx, column in enumerate(matched_df.columns, 1):
            cell = full_matches_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(matched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = full_matches_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, column in enumerate(matched_df.columns, 1):
            full_matches_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(str(column)) + 2, 12)

    # Create Program Mismatches Sheet
    if program_mismatched_df is not None and not program_mismatched_df.empty:
        program_mismatch_sheet = wb.create_sheet("Program Mismatches")
        
        # Add headers
        for col_idx, column in enumerate(program_mismatched_df.columns, 1):
            cell = program_mismatch_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(program_mismatched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = program_mismatch_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, column in enumerate(program_mismatched_df.columns, 1):
            program_mismatch_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(str(column)) + 2, 12)

    # Create Time Matched/Program Mismatched Sheet
    if time_matched_program_mismatched_df is not None and not time_matched_program_mismatched_df.empty:
        time_program_sheet = wb.create_sheet("Time Matched-Program Mismatched")
        
        # Add headers
        for col_idx, column in enumerate(time_matched_program_mismatched_df.columns, 1):
            cell = time_program_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(time_matched_program_mismatched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = time_program_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, column in enumerate(time_matched_program_mismatched_df.columns, 1):
            time_program_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(str(column)) + 2, 12)

    # Create Unmatched TC Entries Sheet
    if unmatched_df is not None and not unmatched_df.empty:
        unmatched_sheet = wb.create_sheet("Unmatched TC Entries")
        
        # Add headers
        for col_idx, column in enumerate(unmatched_df.columns, 1):
            cell = unmatched_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(unmatched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = unmatched_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, column in enumerate(unmatched_df.columns, 1):
            unmatched_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(str(column)) + 2, 12)

    return wb

def highlight_matched_rows_excel(media_watch_filtered, matched_mw_original, highlight_color="#FFFF00"):
 """
 Highlight duplicate rows in media_watch_filtered that are already present in matched_lmrb.
 
 Args:
     media_watch_filtered (pd.DataFrame): The filtered LMRB data.
     matched_lmrb (pd.DataFrame): The matched LMRB data.
     highlight_color (str): Hex color code for highlighting duplicates (default is yellow).
     
 Returns:
     Workbook: An openpyxl Workbook object with highlighted duplicates.
 """
 # Ensure both DataFrames have the same columns for comparison
 common_columns = list(set(media_watch_filtered.columns).intersection(set(matched_mw_original.columns)))
 if not common_columns:
     raise ValueError("No common columns found between media_watch_filtered and matched_lmrb for comparison.")
 
 # Identify duplicate rows
 duplicates_mask = media_watch_filtered[common_columns].apply(tuple, axis=1).isin(
     matched_mw_original[common_columns].apply(tuple, axis=1)
 )
 duplicate_indices = duplicates_mask[duplicates_mask].index.tolist()  # Get indices of duplicates
 
 # Create a new workbook
 wb = Workbook()
 ws = wb.active
 ws.title = "Filtered LMRB Data"

 # Add headers
 for col_idx, column in enumerate(media_watch_filtered.columns, 1):
     cell = ws.cell(row=1, column=col_idx, value=column)
     cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Header background color

 # Add data and highlight duplicates
 for row_idx, row in enumerate(media_watch_filtered.itertuples(index=False), 2):
     for col_idx, value in enumerate(row, 1):
         cell = ws.cell(row=row_idx, column=col_idx, value=value)
         
         # Highlight the entire row if it is a duplicate
         if row_idx - 2 in duplicate_indices:  # Adjusting for zero-based index
             cell.fill = PatternFill(
                 start_color=highlight_color.lstrip('#'),
                 end_color=highlight_color.lstrip('#'),
                 fill_type="solid"
             )

 # Adjust column widths safely
 for col_idx in range(1, ws.max_column + 1):
     column_letter = get_column_letter(col_idx)
     max_length = len(str(media_watch_filtered.columns[col_idx-1])) + 2  # Start with header length
     
     # Check values in each row
     for row_idx in range(2, ws.max_row + 1):
         cell = ws.cell(row=row_idx, column=col_idx)
         if hasattr(cell, 'value') and cell.value:
             max_length = max(max_length, len(str(cell.value)))
     
     ws.column_dimensions[column_letter].width = max_length

 return wb

def create_schedule_compliance_report(matched_df, schedule_df, unmatched_schedule_df):
    """Create a comprehensive Excel report for schedule compliance."""
    wb = Workbook()

    # Create Summary Sheet
    summary_sheet = wb.active
    summary_sheet.title = "Schedule Compliance Summary"

    # Add title and header
    summary_sheet.merge_cells('A1:H1')
    title_cell = summary_sheet['A1']
    title_cell.value = "Media Schedule Compliance Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Add metadata
    summary_sheet['A3'] = "Report Date:"
    summary_sheet['B3'] = pd.Timestamp.now().strftime("%Y-%m-%d")

    # Add match statistics
    summary_sheet['A5'] = "Schedule Compliance Statistics"
    summary_sheet['A5'].font = Font(bold=True)

    summary_sheet['A6'] = "Total Scheduled Spots:"
    summary_sheet['B6'] = len(schedule_df) if schedule_df is not None else 0

    summary_sheet['A7'] = "Verified Aired Spots:"
    summary_sheet['B7'] = len(matched_df) if matched_df is not None else 0

    summary_sheet['A8'] = "Missed Spots:"
    summary_sheet['B8'] = len(unmatched_schedule_df) if unmatched_schedule_df is not None else 0

    # Calculate compliance rate
    if schedule_df is not None and len(schedule_df) > 0:
        compliance_rate = (len(matched_df) / len(schedule_df)) * 100 if matched_df is not None else 0
        summary_sheet['A9'] = "Compliance Rate:"
        summary_sheet['B9'] = f"{compliance_rate:.2f}%"

    # Add theme-based summary
    if schedule_df is not None and matched_df is not None:
        # Group by theme and duration for more detailed compliance analysis
        if 'Schedule_Theme' in matched_df.columns and 'Dur' in matched_df.columns and 'Advt_Theme' in schedule_df.columns and 'Dur' in schedule_df.columns:
            # Group by theme and duration
            matched_theme_counts = matched_df.groupby(['Schedule_Theme', 'Dur']).size().reset_index(name='Aired_Count')
            schedule_theme_counts = schedule_df.groupby(['Advt_Theme', 'Dur']).size().reset_index(name='Scheduled_Count')
            schedule_theme_counts.rename(columns={'Advt_Theme': 'Theme'}, inplace=True)
            matched_theme_counts.rename(columns={'Schedule_Theme': 'Theme'}, inplace=True)
            
            # Merge to get comparison
            theme_summary = pd.merge(schedule_theme_counts, matched_theme_counts, 
                                    on=['Theme', 'Dur'], how='left')
            theme_summary['Aired_Count'].fillna(0, inplace=True)
            theme_summary['Aired_Count'] = theme_summary['Aired_Count'].astype(int)
            theme_summary['Missed_Count'] = theme_summary['Scheduled_Count'] - theme_summary['Aired_Count']
            theme_summary['Compliance_Rate'] = (theme_summary['Aired_Count'] / theme_summary['Scheduled_Count'] * 100).round(2)
            
            # Add to sheet
            summary_sheet['A11'] = "Compliance by Theme and Duration"
            summary_sheet['A11'].font = Font(bold=True)
            
            headers = ['Theme', 'Duration', 'Scheduled', 'Aired', 'Missed', 'Compliance %']
            for col_idx, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=12, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, theme_row in enumerate(theme_summary.itertuples(index=False), 13):
                summary_sheet.cell(row=row_idx, column=1, value=theme_row.Theme)
                summary_sheet.cell(row=row_idx, column=2, value=theme_row.Dur)
                summary_sheet.cell(row=row_idx, column=3, value=theme_row.Scheduled_Count)
                summary_sheet.cell(row=row_idx, column=4, value=theme_row.Aired_Count)
                summary_sheet.cell(row=row_idx, column=5, value=theme_row.Missed_Count)
                summary_sheet.cell(row=row_idx, column=6, value=f"{theme_row.Compliance_Rate}%")
        else:
            # Fallback to just theme-based grouping
            matched_theme_counts = matched_df.groupby('Schedule_Theme').size().reset_index(name='Aired_Count')
            schedule_theme_counts = schedule_df.groupby('Advt_Theme').size().reset_index(name='Scheduled_Count')
            schedule_theme_counts.rename(columns={'Advt_Theme': 'Theme'}, inplace=True)
            matched_theme_counts.rename(columns={'Schedule_Theme': 'Theme'}, inplace=True)
            
            # Merge to get comparison
            theme_summary = pd.merge(schedule_theme_counts, matched_theme_counts, on='Theme', how='left')
            theme_summary['Aired_Count'].fillna(0, inplace=True)
            theme_summary['Aired_Count'] = theme_summary['Aired_Count'].astype(int)
            theme_summary['Missed_Count'] = theme_summary['Scheduled_Count'] - theme_summary['Aired_Count']
            theme_summary['Compliance_Rate'] = (theme_summary['Aired_Count'] / theme_summary['Scheduled_Count'] * 100).round(2)
            
            # Add to sheet
            summary_sheet['A11'] = "Compliance by Theme"
            summary_sheet['A11'].font = Font(bold=True)
            
            headers = ['Theme', 'Scheduled', 'Aired', 'Missed', 'Compliance %']
            for col_idx, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=12, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, theme_row in enumerate(theme_summary.itertuples(index=False), 13):
                summary_sheet.cell(row=row_idx, column=1, value=theme_row.Theme)
                summary_sheet.cell(row=row_idx, column=2, value=theme_row.Scheduled_Count)
                summary_sheet.cell(row=row_idx, column=3, value=theme_row.Aired_Count)
                summary_sheet.cell(row=row_idx, column=4, value=theme_row.Missed_Count)
                summary_sheet.cell(row=row_idx, column=5, value=f"{theme_row.Compliance_Rate}%")

    # Create Matched Schedule Sheet
    if matched_df is not None and not matched_df.empty:
        matched_sheet = wb.create_sheet("Aired as Scheduled")
        
        # Add headers
        important_columns = [
            'Schedule_Theme', 'Media_Watch_Theme', 'Date', 'Schedule_Program', 'Program', 'Dur',
            'Media_Watch_Time', 'Duration_Difference', 'Match_Status'
        ]
        
        # Find available columns
        available_columns = [col for col in important_columns if col in matched_df.columns]
        remaining_columns = [col for col in matched_df.columns if col not in available_columns]
        display_columns = available_columns + remaining_columns
        
        for col_idx, column in enumerate(display_columns, 1):
            if column in matched_df.columns:
                cell = matched_sheet.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(matched_df.itertuples(index=False), 2):
            for col_idx, column in enumerate(display_columns, 1):
                if column in matched_df.columns:
                    col_pos = matched_df.columns.get_loc(column)
                    value = row[col_pos]
                    matched_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        ws = wb.active
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(matched_df.columns[col_idx-1])) + 2  # Start with header length
        #for col_idx, column in enumerate(display_columns, 1):
           # if column in matched_df.columns:
                #matched_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(column) + 2, 15)

    # Create Missed Schedule Sheet
    if unmatched_schedule_df is not None and not unmatched_schedule_df.empty:
        missed_sheet = wb.create_sheet("Missed Schedule Spots")
        
        # Add headers
        important_columns = [
            'Advt_Theme', 'Date', 'Program', 'Advt_time', 'Dur', 'Match_Status'
        ]
        
        # Find available columns
        available_columns = [col for col in important_columns if col in unmatched_schedule_df.columns]
        #remaining_columns = [col for col in unmatched_schedule_df.columns if col not in available_columns]
        display_columns = available_columns #+ #remaining_columns
        
        for col_idx, column in enumerate(display_columns, 1):
            if column in unmatched_schedule_df.columns:
                cell = missed_sheet.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")  # Light red for missed
        
        # Add data
        for row_idx, row in enumerate(unmatched_schedule_df.itertuples(index=False), 2):
            for col_idx, column in enumerate(display_columns, 1):
                if column in unmatched_schedule_df.columns:
                    col_pos = unmatched_schedule_df.columns.get_loc(column)
                    value = row[col_pos]
                    missed_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, column in enumerate(display_columns, 1):
            if column in unmatched_schedule_df.columns:
                missed_sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(column) + 2, 15)

    return wb

def read_uploaded_file(uploaded_file):
    """Read an uploaded file into a pandas DataFrame."""
    if uploaded_file is None:
        return None

    try:
        file_extension = uploaded_file.name.split('.')[-1].lower()
        
        if file_extension == 'xlsx':
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        elif file_extension == 'xls':
            df = pd.read_excel(uploaded_file, engine='xlrd')
        elif file_extension == 'csv':
            df = pd.read_csv(uploaded_file)
        else:
            st.error(f"Unsupported file format: {file_extension}")
            return None
        
        df = df.reset_index(drop=True)
        df = clean_column_names(df)
        
        return df
        
    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        return None

def create_summary_charts(matched_data, tc_data, lmrb_data, schedule_data, channel):
    """Create visual charts for data analysis summary"""
    charts = {}

    try:
        # Helper function to safely convert values to integers
        def safe_int_convert(value):
            try:
                if pd.isna(value) or np.isinf(value):
                    return 0
                return int(float(value))
            except:
                return 0
        
        # Matched vs Unmatched Summary Pie Chart
        if lmrb_data is not None:
            total_lmrb = len(lmrb_data)
            if matched_data is not None:
                matched_count = len(matched_data)
                unmatched_count = total_lmrb - matched_count
                
                labels = ['Matched', 'Unmatched']
                values = [matched_count, unmatched_count]
                colors = ['#4CAF50', '#FFA726']  # Green for matched, orange for unmatched
                
                fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.4, marker_colors=colors)])
                fig.update_layout(
                    title_text=f'LMRB Data Matching Summary - {channel}',
                    title_x=0.5,
                    annotations=[dict(text=f'{matched_count}/{total_lmrb}', x=0.5, y=0.5, font_size=20, showarrow=False)]
                )
                charts['match_summary'] = fig
        
        # Theme Distribution Chart (top 10 themes)
        if matched_data is not None and 'Media_Watch_Theme' in matched_data.columns:
            theme_counts = matched_data['Media_Watch_Theme'].value_counts().nlargest(10).reset_index()
            theme_counts.columns = ['Theme', 'Count']
            
            fig = px.bar(
                theme_counts, 
                x='Theme', 
                y='Count',
                title=f'Top 10 Themes Distribution - {channel}',
                text='Count'
            )
            fig.update_layout(xaxis_title="Theme", yaxis_title="Number of Spots")
            charts['theme_distribution'] = fig
        
        # Schedule Compliance Chart
        if matched_data is not None and schedule_data is not None and 'Schedule_Theme' in matched_data.columns:
            # Group by theme and duration
            if 'Dur' in matched_data.columns and 'Dur' in schedule_data.columns:
                grouped = matched_data.groupby(['Schedule_Theme', 'Dur']).size().reset_index(name='Aired_Count')
                schedule_counts = schedule_data.groupby(['Advt_Theme', 'Dur']).size().reset_index(name='Scheduled_Count')
                schedule_counts.rename(columns={'Advt_Theme': 'Theme'}, inplace=True)
                grouped.rename(columns={'Schedule_Theme': 'Theme'}, inplace=True)
                
                merged = pd.merge(schedule_counts, grouped, on=['Theme', 'Dur'], how='left')
            else:
                # Fallback to just theme
                grouped = matched_data.groupby('Schedule_Theme').size().reset_index(name='Aired_Count')
                schedule_counts = schedule_data.groupby('Advt_Theme').size().reset_index(name='Scheduled_Count')
                schedule_counts.rename(columns={'Advt_Theme': 'Theme'}, inplace=True)
                grouped.rename(columns={'Schedule_Theme': 'Theme'}, inplace=True)
                
                merged = pd.merge(schedule_counts, grouped, on='Theme', how='left')
            
            merged['Aired_Count'].fillna(0, inplace=True)
            merged['Aired_Count'] = merged['Aired_Count'].astype(int)
            merged['Missed_Count'] = merged['Scheduled_Count'] - merged['Aired_Count']
            merged['Missed_Count'] = merged['Missed_Count'].clip(lower=0)  # No negative missed count
            
            # Get top 10 themes by scheduled count
            top_themes = merged.nlargest(10, 'Scheduled_Count')
            
            fig = go.Figure()
            
            fig.add_trace(go.Bar(
                x=top_themes['Theme'],
                y=top_themes['Scheduled_Count'],
                name='Scheduled',
                marker_color='#2196F3'  # Blue
            ))
            
            fig.add_trace(go.Bar(
                x=top_themes['Theme'],
                y=top_themes['Aired_Count'],
                name='Aired',
                marker_color='#4CAF50'  # Green
            ))
            
            fig.add_trace(go.Bar(
                x=top_themes['Theme'],
                y=top_themes['Missed_Count'],
                name='Missed',
                marker_color='#F44336'  # Red
            ))
            
            fig.update_layout(
                barmode='group',
                title_text=f'Schedule Compliance - Top 10 Themes - {channel}',
                xaxis_title="Theme",
                yaxis_title="Count",
                legend_title="Status"
            )
            charts['compliance_chart'] = fig
        
        # Program Distribution Chart
        if matched_data is not None and 'Program' in matched_data.columns:
            program_counts = matched_data['Program'].value_counts().nlargest(10).reset_index()
            program_counts.columns = ['Program', 'Count']
            
            fig = px.bar(
                program_counts,
                x='Program',
                y='Count',
                title=f'Top 10 Programs - {channel}',
                text='Count',
                color='Count',
                color_continuous_scale=px.colors.sequential.Viridis
            )
            fig.update_layout(xaxis_title="Program", yaxis_title="Number of Spots")
            charts['program_distribution'] = fig
        
        # Day/Time Distribution Heatmap - WITH ERROR HANDLING
        if matched_data is not None and 'Dd' in matched_data.columns and 'Advt_time' in matched_data.columns:
            # Extract hour from time with proper error handling
            matched_data_copy = matched_data.copy()
            
            if 'Time_Seconds' in matched_data_copy.columns:
                matched_data_copy['Hour'] = matched_data_copy['Time_Seconds'].fillna(0).apply(
                    lambda x: safe_int_convert(x / 3600)
                )
            else:
                # Handle string format time safely
                try:
                    matched_data_copy['Hour'] = matched_data_copy['Advt_time'].astype(str).str.split(':', expand=True)[0].fillna('0').apply(safe_int_convert)
                except:
                    # Fallback if string splitting fails
                    matched_data_copy['Hour'] = 0
            
            # Make sure Dd column is valid for crosstab
            matched_data_copy['Dd'] = matched_data_copy['Dd'].fillna(0).apply(safe_int_convert)
            
            # Create day-hour crosstab
            hour_day = pd.crosstab(matched_data_copy['Hour'], matched_data_copy['Dd'])
            
            fig = go.Figure(data=go.Heatmap(
                z=hour_day.values,
                x=hour_day.columns,
                y=hour_day.index,
                colorscale='Viridis',
                showscale=True
            ))
            
            fig.update_layout(
                title=f'Spot Distribution by Hour and Day - {channel}',
                xaxis_title='Day of Month',
                yaxis_title='Hour of Day',
                height=500
            )
            charts['time_distribution'] = fig
            
    except Exception as e:
        st.warning(f"Could not generate all charts due to error: {str(e)}")

    return charts

def main():
    st.set_page_config(page_title="Media Reconciliation Tool", layout="wide")

    # Custom CSS for better UI appearance
    st.markdown("""
    <style>
    .stSelectbox label, .stButton button {
        font-weight: bold;
    }
    .dataframe {
        font-size: 12px;
    }
    .stTabs [data-baseweb="tab"] {
        font-weight: bold;
    }
    </style>
    """, unsafe_allow_html=True)

    st.title("Media Data Reconciliation Tool")
    st.markdown("""
    Upload data files, select channel, map themes, and find matching records between LMRB, TC and Schedule. 
    The tool provides comprehensive matching and analytics for media data reconciliation.
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.header("Upload Files")
        media_watch_file = st.file_uploader("Upload LMRB Data", type=["xlsx", "xls", "csv"])
        tc_file = st.file_uploader("Upload TC Data", type=["xlsx", "xls", "csv"])
        schedule_file = st.file_uploader("Upload Schedule Data", type=["xlsx", "xls", "csv"])
        prev_month_file = st.file_uploader("Upload Previous Month LMRB Data (optional)", type=["xlsx", "xls", "csv"])
        
        st.header("Matching Configuration")
        time_tolerance = st.slider("Time Matching Tolerance (seconds)", min_value=5, max_value=60, value=30, step=5, 
                                help="Maximum time difference (in seconds) allowed for matching times between datasets")
        
        st.markdown("---")
        st.markdown("### About")
        st.info("This tool helps media analysts reconcile LMRB, TC, and Schedule data for improved reporting accuracy.")

    # Initialize session state variables
    if 'mw_tc_matched_data' not in st.session_state:
        st.session_state.mw_tc_matched_data = None

    if 'mw_schedule_matched_data' not in st.session_state:
        st.session_state.mw_schedule_matched_data = None

    if 'final_matched_data' not in st.session_state:
        st.session_state.final_matched_data = None

    if 'program_mismatched_data' not in st.session_state:
        st.session_state.program_mismatched_data = None

    if 'theme_mapping' not in st.session_state:
        st.session_state.theme_mapping = []

    if 'matched_mw_indices' not in st.session_state:
        st.session_state.matched_mw_indices = []

    if 'unmatched_schedule_data' not in st.session_state:
        st.session_state.unmatched_schedule_data = None

    # Initialize manual matching state variables
    if 'manual_matched_indices' not in st.session_state:
        st.session_state.manual_matched_indices = []

    if 'selected_tc_rows' not in st.session_state:
        st.session_state.selected_tc_rows = []

    if 'selected_lmrb_rows' not in st.session_state:
        st.session_state.selected_lmrb_rows = []

    if 'tc_lmrb_manual_matches' not in st.session_state:
        st.session_state.tc_lmrb_manual_matches = []
        
    if 'selected_tc_theme' not in st.session_state:
        st.session_state.selected_tc_theme = None

    if 'selected_schedule_theme' not in st.session_state:
        st.session_state.selected_schedule_theme = None

    # Read uploaded files
    media_watch_df = read_uploaded_file(media_watch_file)
    tc_df = read_uploaded_file(tc_file)
    schedule_df = read_uploaded_file(schedule_file)
    prev_month_df = read_uploaded_file(prev_month_file)

    if media_watch_df is not None:
        st.success("LMRB data file loaded successfully!")
        
        if "Channel" in media_watch_df.columns:
            available_channels = sorted(media_watch_df["Channel"].unique())
            channel = st.selectbox("Select Channel", options=available_channels)
            
            media_watch_filtered = filter_by_channel(media_watch_df, channel)
            
            if prev_month_df is not None:
                prev_month_filtered = filter_by_channel(prev_month_df, channel)
                cleaned_mw_df, duplicate_count = remove_duplicates(media_watch_filtered, prev_month_filtered)
                
                st.info(f"Removed {duplicate_count} duplicate records from previous month's data.")
                media_watch_filtered = cleaned_mw_df
            
            st.subheader("Theme Mapping")
            
           # media_watch_theme_col = st.selectbox(
                #"Select LMRB Theme Column", 
                #options=[col for col in media_watch_filtered.columns if 'theme' in col.lower() or 'aadvt' in col.lower()],
                #key="mw_theme_col"
            #)
            
            # Automatically detect the LMRB Theme Column
            media_watch_theme_col = next(
            (col for col in media_watch_filtered.columns if 'theme' in col.lower() or 'aadvt' in col.lower()), 
            None  # Default to None if no matching column is found
            )

            if not media_watch_theme_col:
                st.error("No column matching 'theme' or 'aadvt' found in the Media Watch data.")
            else:

                media_watch_std = standardize_dataframe(
                    media_watch_filtered,
                    theme_col=media_watch_theme_col,
                    program_col=next((col for col in media_watch_filtered.columns if 'program' in col.lower()), None),
                    time_col=next((col for col in media_watch_filtered.columns if 'time' in col.lower() and 'advt' in col.lower()), None),
                    date_col=next((col for col in media_watch_filtered.columns if 'date' in col.lower()), None)
                )
                
                if tc_df is not None:

                    # Automatically detect the LMRB Theme Column
                    tc_theme_col = next(
                    (col for col in tc_df.columns if 'theme' in col.lower() or 'aadvt' in col.lower()), 
                    None  # Default to None if no matching column is found
                    )

                    if not media_watch_theme_col:
                        st.error("No column matching 'theme' or 'aadvt' found in the TC.")
                    else:
                    
                        tc_std = standardize_dataframe(
                            tc_df,
                            theme_col=tc_theme_col,
                            program_col=next((col for col in tc_df.columns if 'program' in col.lower()), None),
                            time_col=next((col for col in tc_df.columns if 'time' in col.lower() and 'spot' in col.lower() or 'air' in col.lower()), None),
                            date_col=next((col for col in tc_df.columns if 'date' in col.lower()), None)
                        )
                        
                    
                
                if schedule_df is not None:
                    
                    # Automatically detect the LMRB Theme Column
                    schedule_theme_col = next(
                    (col for col in schedule_df.columns if 'theme' in col.lower() or 'aadvt' in col.lower()), 
                    None  # Default to None if no matching column is found
                    )

                    if not media_watch_theme_col:
                        st.error("No column matching 'theme' or 'aadvt' found in the TC.")
                    else:
                    
                        schedule_std = standardize_dataframe(
                            schedule_df,
                            theme_col=schedule_theme_col,
                            program_col=next((col for col in schedule_df.columns if 'program' in col.lower()), None),
                            time_col=next((col for col in schedule_df.columns if 'time' in col.lower()), None),
                            date_col=next((col for col in schedule_df.columns if 'date' in col.lower()), None)
                        )
                    
            
            # Create three columns for theme mapping
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.subheader("LMRB Theme")
                media_watch_themes = sorted(media_watch_std['Advt_Theme'].dropna().unique())
                selected_mw_theme = st.selectbox("Select LMRB Theme", options=[""] + media_watch_themes)
            
            with col2:
                st.subheader("TC Theme")
                if tc_std is not None:
                    tc_themes = sorted(tc_std['Advt_Theme'].dropna().unique())
                    selected_tc_theme = st.selectbox("Select TC Theme (leave empty if not in TC)", options=[""] + tc_themes)
                else:
                    selected_tc_theme = ""
                    st.info("No TC data uploaded")
            
            with col3:
                st.subheader("Schedule Theme")
                if schedule_std is not None:
                    schedule_themes = sorted(schedule_std['Advt_Theme'].dropna().unique())
                    selected_schedule_theme = st.selectbox("Select Schedule Theme", options=[""] + schedule_themes)
                    
                    # Duration selection for schedule theme
                    st.write("Duration (seconds)")
                    # Common durations for quick selection
                    duration_options = []
                    if 'Dur' in schedule_std.columns:
                        duration_options = sorted(schedule_std['Dur'].dropna().unique())
                    elif 'Duration' in schedule_std.columns:
                        duration_options = sorted(schedule_std['Duration'].dropna().unique())
                    
                    col3_1, col3_2 = st.columns([2, 1])
                    with col3_1:
                        selected_duration = st.selectbox("Duration Presets", options=[""] + [str(x) for x in duration_options], key="duration_preset")
                    with col3_2:
                        custom_duration = st.number_input("Custom", min_value=0, max_value=300, value=0, step=5, key="custom_duration")
                    
                    # Use either the preset or custom duration
                    selected_duration = custom_duration if custom_duration > 0 else (int(selected_duration) if selected_duration else None)
                else:
                    selected_schedule_theme = ""
                    selected_duration = None
                    st.info("No Schedule data uploaded")
            
            if st.button("Add Theme Mapping", key="add_theme_mapping"):
                if selected_mw_theme:
                    is_duplicate = False
                    for mapping in st.session_state.theme_mapping:
                        if normalize_theme_name(mapping['media_watch_theme']) == normalize_theme_name(selected_mw_theme):
                            is_duplicate = True
                            break
                    
                    if not is_duplicate:
                        mapping_entry = {
                            'media_watch_theme': selected_mw_theme,
                        }
                        
                        if selected_tc_theme:
                            mapping_entry['tc_theme'] = selected_tc_theme
                        else:
                            mapping_entry['tc_theme'] = ""
                            
                        if selected_schedule_theme:
                            mapping_entry['schedule_theme'] = selected_schedule_theme
                            # Add duration to the mapping if available
                            if selected_duration:
                                try:
                                    mapping_entry['duration'] = int(selected_duration)
                                except ValueError:
                                    mapping_entry['duration'] = None
                        
                        st.session_state.theme_mapping.append(mapping_entry)
                        
                        mapping_msg = f"Added mapping: {selected_mw_theme}"
                        if selected_tc_theme:
                            mapping_msg += f" → TC: {selected_tc_theme}"
                        if selected_schedule_theme:
                            mapping_msg += f" → Schedule: {selected_schedule_theme}"
                            if selected_duration:
                                mapping_msg += f" ({selected_duration}s)"
                            
                        st.success(mapping_msg)
                    else:
                        st.warning(f"A mapping for '{selected_mw_theme}' already exists.")
                else:
                    st.warning("Please select at least an LMRB theme to create a mapping.")
            
            if st.session_state.theme_mapping:
                st.subheader("Current Theme Mappings")
                
                # Create DataFrame from theme mappings for better display
                mapping_df = pd.DataFrame(st.session_state.theme_mapping)
                
                # Define column order
                columns_to_show = ['media_watch_theme', 'tc_theme', 'schedule_theme', 'duration']
                display_cols = [col for col in columns_to_show if col in mapping_df.columns]
                
                # Display mappings in styled dataframe
                if not mapping_df.empty:
                    st.dataframe(mapping_df[display_cols], height=150)
                
                # Add delete button
                mapping_to_delete = st.selectbox(
                    "Select theme mapping to delete:", 
                    options=[""] + mapping_df['media_watch_theme'].tolist()
                )
                
                if mapping_to_delete and st.button("Delete Selected Mapping"):
                    for i, mapping in enumerate(st.session_state.theme_mapping):
                        if mapping['media_watch_theme'] == mapping_to_delete:
                            st.session_state.theme_mapping.pop(i)
                            st.success(f"Deleted mapping for '{mapping_to_delete}'")
                            st.rerun()
            
            st.markdown("---")
            st.subheader("Matching Options")
            ignore_date = st.checkbox("Ignore Date (Match Only by Theme, Program, and Duration)", value=False)
            
            match_tab, manual_tab, analytics_tab = st.tabs(["Automatic Matching", "Manual Matching", "Analytics"])
            
            with match_tab:
                st.subheader("Automatic Data Matching")
                
                if st.button("Run Automatic Matching", key="run_auto_match"):
                    if not st.session_state.theme_mapping:
                        st.warning("Please create theme mappings before matching data.")
                    else:
                        all_results = []
                        unmatched_mw_indices = list(range(len(media_watch_std)))
                        
                        with st.spinner("Matching LMRB with TC data..."):
                            if tc_std is not None:
                                try:
                                    matched_tc_results, matched_tc_indices, program_mismatched_df = match_media_watch_with_tc(
                                        media_watch_std, 
                                        tc_std, 
                                        st.session_state.theme_mapping,
                                        ignore_date=ignore_date,
                                        time_tolerance=time_tolerance
                                    )
                                    
                                    if not matched_tc_results.empty:
                                        st.session_state.mw_tc_matched_data = matched_tc_results
                                        st.session_state.matched_mw_indices = matched_tc_indices
                                        st.session_state.program_mismatched_data = program_mismatched_df
                                        
                                        st.success(f"Found {len(matched_tc_results)} matching records between LMRB and TC data!")
                                        if not program_mismatched_df.empty:
                                            st.info(f"Found {len(program_mismatched_df)} records with time matched but program mismatched.")
                                        
                                        # Generate comprehensive TC vs LMRB report
                                        tc_lmrb_report = create_tc_vs_lmrb_excel_report(
                                            matched_tc_results,
                                            program_mismatched_df,
                                            pd.DataFrame(),  # Unmatched (we don't have this yet)
                                            program_mismatched_df,  # Time matched/program mismatched
                                            media_watch_filtered,
                                            f"{channel}_TC_vs_LMRB_report.xlsx"
                                        )
                                        
                                        # Convert report to binary for download
                                        tc_lmrb_buffer = BytesIO()
                                        tc_lmrb_report.save(tc_lmrb_buffer)
                                        tc_lmrb_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="Download TC vs LMRB Reconciliation Report",
                                            data=tc_lmrb_buffer,
                                            file_name=f"{channel}_TC_vs_LMRB_report.xlsx",
                                            mime="application/vnd.ms-excel",
                                            key="download_tc_lmrb_report"
                                        )
                                        
                                        matched_mw_original = media_watch_filtered.iloc[matched_tc_indices].copy()
                                        
                                        excel_buffer = BytesIO()
                                        matched_mw_original.to_excel(excel_buffer, index=False)
                                        excel_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="Download Matched LMRB Data (TC Matched)",
                                            data=excel_buffer,
                                            file_name=f"{channel}_lmrb_tc_matched.xlsx",
                                            mime="application/vnd.ms-excel",
                                            key="download_lmrb_matched"
                                        )
                                        
                                        all_results.append(matched_tc_results)
                                        
                                        for idx in matched_tc_indices:
                                            if idx in unmatched_mw_indices:
                                                unmatched_mw_indices.remove(idx)
                                except Exception as e:
                                    st.error(f"Error during TC matching: {str(e)}")
                        
                        if st.session_state.mw_tc_matched_data is not None and not st.session_state.mw_tc_matched_data.empty and schedule_std is not None:
                            with st.spinner("Matching TC results with Schedule data..."):
                                try:
                                    schedule_matched_results, schedule_unmatched_results = match_with_schedule(
                                        st.session_state.mw_tc_matched_data,
                                        schedule_std,
                                        st.session_state.theme_mapping,
                                        ignore_date=ignore_date
                                    )
                                    
                                    if not schedule_matched_results.empty:
                                        st.success(f"Found {len(schedule_matched_results)} matching records between TC results and Schedule data!")
                                        
                                        # Extract matched schedule themes and dates for finding unmatched spots
                                        matched_schedule_themes = {}
                                        
                                        for _, row in schedule_matched_results.iterrows():
                                            if 'Schedule_Theme' in row:
                                                theme = normalize_theme_name(row['Schedule_Theme'])
                                                date_tuple = (row['Dd'], row['Mn'], row['Yr'])
                                                
                                                if theme not in matched_schedule_themes:
                                                    matched_schedule_themes[theme] = []
                                                    
                                                matched_schedule_themes[theme].append(date_tuple)
                                                
                                        # Find unmatched schedule spots
                                        unmatched_schedule_df = find_unmatched_schedule_spots(
                                            schedule_std, 
                                            matched_schedule_themes
                                        )
                                        
                                        # Store unmatched schedule data for manual matching
                                        st.session_state.unmatched_schedule_data = unmatched_schedule_df
                                        
                                        # Create schedule compliance report
                                        schedule_report = create_schedule_compliance_report(
                                            schedule_matched_results,
                                            schedule_std,
                                            unmatched_schedule_df
                                        )
                                        
                                        # Convert report to binary for download
                                        schedule_buffer = BytesIO()
                                        schedule_report.save(schedule_buffer)
                                        schedule_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="Download Schedule Compliance Report",
                                            data=schedule_buffer,
                                            file_name=f"{channel}_schedule_compliance_report.xlsx",
                                            mime="application/vnd.ms-excel",
                                            key="download_schedule_report"
                                        )
                                        
                                        all_results.append(schedule_matched_results)
                                except Exception as e:
                                    st.error(f"Error during Schedule matching: {str(e)}")
                        
                        if schedule_std is not None and unmatched_mw_indices:
                            with st.spinner("Matching unmatched LMRB data directly with Schedule..."):
                                try:
                                    unmatched_rows = []
                                    for idx in unmatched_mw_indices:
                                        if idx < len(media_watch_std):
                                            unmatched_rows.append(media_watch_std.iloc[idx])
                                    
                                    if unmatched_rows:
                                        unmatched_mw_df = pd.DataFrame(unmatched_rows).reset_index(drop=True)
                                        
                                        direct_schedule_results, direct_schedule_indices = match_media_watch_with_schedule(
                                            unmatched_mw_df,
                                            schedule_std,
                                            st.session_state.theme_mapping,
                                            ignore_date=ignore_date
                                        )
                                        
                                        if not direct_schedule_results.empty:
                                            st.success(f"Found {len(direct_schedule_results)} direct matches between LMRB and Schedule data!")
                                            
                                            direct_original_indices = [unmatched_mw_indices[i] for i in direct_schedule_indices if i < len(unmatched_mw_indices)]
                                            if direct_original_indices:
                                                direct_mw_original = media_watch_filtered.iloc[direct_original_indices].copy()
                                                
                                                excel_buffer = BytesIO()
                                                direct_mw_original.to_excel(excel_buffer, index=False)
                                                excel_buffer.seek(0)
                                                
                                                st.download_button(
                                                    label="Download Original LMRB Data (Schedule Direct Matched)",
                                                    data=excel_buffer,
                                                    file_name=f"{channel}_original_lmrb_schedule_matched.xlsx",
                                                    mime="application/vnd.ms-excel",
                                                    key="download_lmrb_schedule_matched"
                                                )
                                            
                                            all_results.append(direct_schedule_results)
                                            
                                            # Update matched indices
                                            for idx in direct_original_indices:
                                                if idx in unmatched_mw_indices:
                                                    unmatched_mw_indices.remove(idx)
                                                    st.session_state.matched_mw_indices.append(idx)
                                except Exception as e:
                                    st.error(f"Error during direct Schedule matching: {str(e)}")
                        
                        if all_results:
                            for i in range(len(all_results)):
                                all_results[i] = all_results[i].reset_index(drop=True)
                                
                            final_results = pd.concat(all_results, ignore_index=True)
                            st.session_state.final_matched_data = final_results
                            
                            st.subheader("Automatic Matching Results")
                            st.dataframe(final_results, height=300)
                            
                            # Generate comprehensive summary report
                            summary = generate_comprehensive_summary_report(
                                final_results, 
                                media_watch_filtered, 
                                schedule_std
                            )
                            
                            if summary is not None:
                                st.subheader("Reconciliation Summary")
                                st.dataframe(summary, height=300)
                                
                                # Export summary
                                summary_buffer = BytesIO()
                                summary.to_excel(summary_buffer, index=False)
                                summary_buffer.seek(0)
                                
                                st.download_button(
                                    label="Download Reconciliation Summary",
                                    data=summary_buffer,
                                    file_name=f"{channel}_reconciliation_summary.xlsx",
                                    mime="application/vnd.ms-excel",
                                    key="download_summary"
                                )
                            
                            # Export final results with highlighting
                            highlight_color = st.color_picker("Choose highlight color for Excel", "#FFFF00")
                            
                            excel_buffer = BytesIO()
                            wb = highlight_matched_rows_excel(media_watch_filtered, matched_mw_original, highlight_color)
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            st.download_button(
                                label="Download Combined Matching Results",
                                data=excel_buffer,
                                file_name=f"{channel}_combined_matching_results.xlsx",
                                mime="application/vnd.ms-excel",
                                key="download_combined_results"
                            )
                            
                            # Create charts for analytics tab
                            charts = create_summary_charts(
                                final_results, 
                                tc_std, 
                                media_watch_filtered, 
                                schedule_std,
                                channel
                            )
                            
                            # Create PDF summary option
                            if st.button("Generate PDF Report", key="generate_pdf"):
                                with st.spinner("Creating PDF report... This may take a moment."):
                                    try:
                                        # Call the tables-only version that doesn't use images
                                        pdf_filename = create_summary_pdf_tables_only(
                                            summary, 
                                            channel,
                                            st.session_state.final_matched_data,
                                            schedule_std,
                                            st.session_state.unmatched_schedule_data
                                        )
                                        
                                        # Read the generated PDF
                                        with open(pdf_filename, "rb") as f:
                                            pdf_bytes = f.read()
                                        
                                        st.success("PDF Summary created successfully!")
                                        st.download_button(
                                            label="Download PDF Summary Report",
                                            data=pdf_bytes,
                                            file_name=pdf_filename,
                                            mime="application/pdf",
                                            key="download_pdf"
                                        )
                                        
                                        # Clean up the PDF file after download
                                        try:
                                            os.remove(pdf_filename)
                                        except:
                                            pass
                                            
                                        # Also offer Excel as an alternative
                                        excel_buffer = BytesIO()
                                        summary.to_excel(excel_buffer, index=False, sheet_name="Summary")
                                        excel_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="Download Excel Summary (with formatting)",
                                            data=excel_buffer,
                                            file_name=f"{channel}_reconciliation_summary.xlsx",
                                            mime="application/vnd.ms-excel",
                                            key="download_summary_excel"
                                        )
                                        
                                    except Exception as e:
                                        st.error(f"Error creating PDF: {str(e)}")
                        else:
                            st.warning("No matches found in any of the matching processes.")
            
            with manual_tab:
                st.subheader("Manual Matching Interface")
                
                if media_watch_filtered is not None:
                    # Get all matched indices from both automatic and manual methods
                    all_matched_indices = set()
                    if hasattr(st.session_state, 'matched_mw_indices'):
                        all_matched_indices.update(st.session_state.matched_mw_indices)
                    all_matched_indices.update(st.session_state.manual_matched_indices)
                    
                    # Convert to list and sort
                    all_matched_indices = sorted(list(all_matched_indices))
                    
                    # Get matched and unmatched data from LMRB
                    matched_lmrb = media_watch_filtered.iloc[all_matched_indices].reset_index(drop=True) if all_matched_indices else pd.DataFrame()
                    
                    # Get unmatched LMRB data
                    all_indices = set(range(len(media_watch_filtered)))
                    unmatched_lmrb_indices = sorted(list(all_indices - set(all_matched_indices)))
                    unmatched_lmrb = media_watch_filtered.iloc[unmatched_lmrb_indices].reset_index(drop=True) if unmatched_lmrb_indices else pd.DataFrame()
                    
                    # Get TC data that wasn't matched
                    matched_tc_themes = set()
                    if st.session_state.final_matched_data is not None and not st.session_state.final_matched_data.empty and 'TC_Theme' in st.session_state.final_matched_data.columns:
                        matched_tc_themes.update(st.session_state.final_matched_data['TC_Theme'].unique())
                    
                    # Manually matched TC themes
                    for match in st.session_state.tc_lmrb_manual_matches:
                        if 'tc_theme' in match:
                            matched_tc_themes.add(match['tc_theme'])
                    
                    # Filter to get unmatched TC data
                    if tc_std is not None:
                        unmatched_tc = tc_std[~tc_std['Advt_Theme'].isin(matched_tc_themes)].reset_index(drop=True)
                    else:
                        unmatched_tc = pd.DataFrame()
                            
                    # Display match status counts
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total LMRB Records", len(media_watch_filtered))
                    with col2:
                        st.metric("Matched Records", len(matched_lmrb))
                    with col3:
                        st.metric("Unmatched LMRB", len(unmatched_lmrb))
                    with col4:
                        if tc_std is not None:
                            st.metric("Unmatched TC", len(unmatched_tc))
                    
                    # Display matched data
                    st.markdown("### Matched Data")
                    if not matched_lmrb.empty:
                        # Combine automatic and manual matches
                        st.write(f"Total matched entries: {len(matched_lmrb)}")
                        
                        # Add info about match source
                        match_source = []
                        for idx in all_matched_indices:
                            if idx in st.session_state.manual_matched_indices:
                                match_source.append("Manual")
                            else:
                                match_source.append("Automatic")
                                
                        # Display matched data with essential columns
                        display_cols = ['Advt_Theme', 'Program', 'Date', 'Advt_time']
                        display_cols = [c for c in display_cols if c in matched_lmrb.columns]
                        
                        # Add any remaining columns if essential ones aren't available
                        if len(display_cols) < 3 and len(matched_lmrb.columns) >= 3:
                            for col in matched_lmrb.columns[:3]:
                                if col not in display_cols:
                                    display_cols.append(col)
                        
                        matched_display = matched_lmrb[display_cols].copy()
                        matched_display['Match Source'] = match_source
                        
                        st.dataframe(matched_display, height=200)
                        
                        # Download matched data
                        excel_buffer = BytesIO()
                        matched_lmrb.to_excel(excel_buffer, index=False)
                        excel_buffer.seek(0)
                        
                        st.download_button(
                            label="Download Final Matched Data",
                            data=excel_buffer,
                            file_name=f"{channel}_final_matched_data.xlsx",
                            mime="application/vnd.ms-excel",
                            key="download_final_matched"
                        )
                    else:
                        st.info("No matched data yet. Use the tools below to match data manually.")
                    
                    st.markdown("---")
                    
                    # Tabs for TC matching and Schedule matching
                    tc_tab, schedule_tab = st.tabs(["TC Data Matching", "Schedule Data Matching"])
                    
                    with tc_tab:
                        # Unmatched TC Data section
                        st.markdown("### Unmatched TC Data")
                        
                        # TC Theme selection for filtering
                        if not unmatched_tc.empty:
                            # Group unmatched TC themes
                            tc_themes = sorted(unmatched_tc['Advt_Theme'].unique())
                            
                            selected_tc_theme = st.selectbox(
                                "Select TC Theme to Match",
                                options=[""] + tc_themes,
                                key="select_tc_theme_for_manual"
                            )
                            
                            if selected_tc_theme:
                                st.session_state.selected_tc_theme = selected_tc_theme
                                
                                # Filter TC data by selected theme
                                filtered_tc = unmatched_tc[unmatched_tc['Advt_Theme'] == selected_tc_theme].copy()
                                
                                if not filtered_tc.empty:
                                    # Simplify display for TC data
                                    display_cols = ['Advt_Theme', 'Date', 'Advt_time', 'Program', 'Dur']
                                    display_cols = [c for c in display_cols if c in filtered_tc.columns]
                                    
                                    st.write(f"TC entries for theme '{selected_tc_theme}':")
                                    
                                    # Prepare TC data for display with row selection
                                    filtered_tc_display = filtered_tc[display_cols].copy()
                                    filtered_tc_display.insert(0, 'Select', False)
                                    
                                    # Format date if it exists
                                    if 'Date' in filtered_tc_display.columns and 'Dd' in filtered_tc.columns and 'Mn' in filtered_tc.columns and 'Yr' in filtered_tc.columns:
                                        filtered_tc_display['Date'] = filtered_tc.apply(
                                            lambda x: f"{x['Dd']}/{x['Mn']}/{x['Yr']}", 
                                            axis=1
                                        )
                                    
                                    # Add row numbers for identification
                                    filtered_tc_display.insert(0, 'TC_Row', range(len(filtered_tc_display)))
                                    
                                    # Mark already selected rows
                                    for idx, row in enumerate(filtered_tc_display['TC_Row']):
                                        for match in st.session_state.tc_lmrb_manual_matches:
                                            if 'tc_row' in match and match['tc_row'] == row and match['tc_theme'] == selected_tc_theme:
                                                filtered_tc_display.loc[idx, 'Select'] = True
                                    
                                    # Display with row selection
                                    st.dataframe(filtered_tc_display, height=200)
                                    
                                    # Allow selection of TC rows with checkboxes
                                    st.write("Select TC entries to match:")
                                    
                                    # Create columns for better layout of checkboxes
                                    checkbox_cols = st.columns(3)
                                    selected_tc_rows = []
                                    
                                    for idx, row in enumerate(filtered_tc_display.itertuples()):
                                        col_idx = idx % 3
                                        checkbox_key = f"tc_row_{row.TC_Row}"
                                        label = f"Row {row.TC_Row}: {row.Date} - {row.Advt_time}"
                                        
                                        if checkbox_cols[col_idx].checkbox(label, key=checkbox_key, value=row.Select):
                                            selected_tc_rows.append({
                                                'row': row.TC_Row,
                                                'theme': selected_tc_theme,
                                                'data': filtered_tc.iloc[idx].to_dict()
                                            })
                                    
                                    # Update session state with selected TC rows
                                    if selected_tc_rows:
                                        st.session_state.selected_tc_rows = selected_tc_rows
                                    else:
                                        st.session_state.selected_tc_rows = []
                                else:
                                    st.info(f"No unmatched TC entries found for theme '{selected_tc_theme}'")
                        else:
                            st.info("No unmatched TC data available.")
                            st.session_state.selected_tc_theme = None
                        
                        st.markdown("---")
                        
                        # Unmatched LMRB Data section with theme filtering
                        st.markdown("### Unmatched LMRB Data")
                        
                        if not unmatched_lmrb.empty:
                            # LMRB theme filter dropdown
                            lmrb_themes = sorted(unmatched_lmrb['Advt_Theme'].unique())
                            selected_filter_theme = st.selectbox(
                                "Filter LMRB entries by theme",
                                options=["All Themes"] + lmrb_themes,
                                key="lmrb_theme_filter"
                            )
                            
                            # Filter LMRB data by selected theme or by TC mapping
                            filtered_lmrb = unmatched_lmrb
                            
                            if selected_filter_theme != "All Themes":
                                filtered_lmrb = unmatched_lmrb[unmatched_lmrb['Advt_Theme'] == selected_filter_theme].copy()
                                st.write(f"LMRB entries for theme '{selected_filter_theme}':")
                            elif st.session_state.selected_tc_theme:
                                # Find LMRB themes mapped to selected TC theme
                                related_lmrb_themes = []
                                for mapping in st.session_state.theme_mapping:
                                    if 'tc_theme' in mapping and mapping['tc_theme'] == st.session_state.selected_tc_theme:
                                        related_lmrb_themes.append(mapping['media_watch_theme'])
                                
                                if related_lmrb_themes:
                                    filtered_lmrb = unmatched_lmrb[unmatched_lmrb['Advt_Theme'].isin(related_lmrb_themes)].copy()
                                    st.write(f"LMRB entries mapped to TC theme '{st.session_state.selected_tc_theme}':")
                            
                            # LMRB data table with improved details and UI
                            if len(filtered_lmrb) > 0:
                                # Display columns for LMRB data
                                display_cols = ['Advt_Theme', 'Date', 'Advt_time', 'Program', 'Dur']
                                available_cols = [col for col in display_cols if col in filtered_lmrb.columns]
                                
                                # Format date if needed
                                if 'Date' not in filtered_lmrb.columns and all(col in filtered_lmrb.columns for col in ['Dd', 'Mn', 'Yr']):
                                    filtered_lmrb['Date'] = filtered_lmrb.apply(
                                        lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                                        axis=1
                                    )
                                    available_cols.append('Date')
                                
                                # Add row identification
                                filtered_lmrb_display = filtered_lmrb[available_cols].copy()
                                filtered_lmrb_display.insert(0, 'LMRB_Row', [unmatched_lmrb_indices[i] for i in range(len(filtered_lmrb))])
                                
                                # Show the filtered LMRB data
                                st.dataframe(filtered_lmrb_display, height=250)
                                
                                # Add search filter
                                search_term = st.text_input("Search in LMRB data (program name, date, etc.)", key="lmrb_search")
                                
                                if search_term:
                                    # Search across all string columns
                                    mask = pd.Series(False, index=range(len(filtered_lmrb_display)))
                                    for col in filtered_lmrb_display.columns:
                                        if filtered_lmrb_display[col].dtype == 'object':
                                            mask = mask | filtered_lmrb_display[col].astype(str).str.contains(search_term, case=False, na=False)
                                    
                                    search_results = filtered_lmrb_display[mask].copy()
                                    if not search_results.empty:
                                        st.write(f"Found {len(search_results)} matching entries:")
                                        st.dataframe(search_results, height=150)
                                        
                                        # Use search results for selection
                                        filtered_lmrb_display = search_results
                                
                                # Selection UI with pagination for large datasets
                                if len(filtered_lmrb_display) > 50:
                                    items_per_page = 50
                                    page = st.number_input("Page", min_value=1, 
                                                        max_value=(len(filtered_lmrb_display) // items_per_page) + 1, 
                                                        value=1)
                                    
                                    start_idx = (page - 1) * items_per_page
                                    end_idx = min(start_idx + items_per_page, len(filtered_lmrb_display))
                                    
                                    st.write(f"Showing entries {start_idx+1} to {end_idx} of {len(filtered_lmrb_display)}")
                                    page_data = filtered_lmrb_display.iloc[start_idx:end_idx]
                                else:
                                    page_data = filtered_lmrb_display
                                
                                # Selection mechanism with checkboxes
                                st.write("Select LMRB entries to match with TC entries:")
                                
                                # Create columns layout for checkboxes
                                selection_cols = st.columns(3)
                                selected_lmrb_rows = []
                                
                                for idx, row in enumerate(page_data.itertuples()):
                                    col_idx = idx % 3
                                    row_id = row.LMRB_Row
                                    
                                    # Create a checkbox with detailed label
                                    label = f"Row {row_id}: {getattr(row, 'Advt_Theme', '')} - "
                                    label += f"{getattr(row, 'Date', '') if hasattr(row, 'Date') else ''} - "
                                    label += f"{getattr(row, 'Advt_time', '') if hasattr(row, 'Advt_time') else ''}"
                                    
                                    if selection_cols[col_idx].checkbox(label, key=f"lmrb_row_{row_id}"):
                                        selected_lmrb_rows.append(row_id)
                                
                                # Update selected rows
                                if selected_lmrb_rows:
                                        st.session_state.selected_lmrb_rows = selected_lmrb_rows

                                        # Show details of selected entries
                                        st.subheader("Selected LMRB Entries:")
                                        selected_entries_df = media_watch_filtered.loc[selected_lmrb_rows]

                                        # Make sure we only use columns that exist in the selected entries
                                        safe_cols = [col for col in available_cols if col in selected_entries_df.columns]

                                        # Format date in selected entries if needed
                                        if 'Date' not in selected_entries_df.columns and all(col in selected_entries_df.columns for col in ['Dd', 'Mn', 'Yr']):
                                            selected_entries_df = selected_entries_df.copy()
                                            selected_entries_df['Date'] = selected_entries_df.apply(
                                                lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                                                axis=1
                                            )
                                            if 'Date' not in safe_cols:
                                                safe_cols.append('Date')

                                        st.dataframe(selected_entries_df[safe_cols], height=150)
                                else:
                                    st.session_state.selected_lmrb_rows = []
                                
                                # Add match button
                                if st.button("Add Manual Matches", key="add_manual_match"):
                                    for tc_item in st.session_state.selected_tc_rows:
                                        tc_row = tc_item['row']
                                        tc_theme = tc_item['theme']
                                        
                                        for lmrb_row in st.session_state.selected_lmrb_rows:
                                            # Check if this match already exists
                                            match_exists = False
                                            for match in st.session_state.tc_lmrb_manual_matches:
                                                if match.get('tc_row') == tc_row and match.get('lmrb_row') == lmrb_row:
                                                    match_exists = True
                                                    break
                                            
                                            if not match_exists:
                                                # Get the LMRB data for this row
                                                lmrb_data = media_watch_filtered.iloc[lmrb_row].to_dict() if lmrb_row < len(media_watch_filtered) else {}
                                                
                                                # Find the mapping for this TC theme to get the correct LMRB theme
                                                mapped_lmrb_theme = None
                                                for mapping in st.session_state.theme_mapping:
                                                    if 'tc_theme' in mapping and mapping['tc_theme'] == tc_theme:
                                                        mapped_lmrb_theme = mapping['media_watch_theme']
                                                        break
                                                
                                                # Add to manual matches with the correct theme
                                                match_entry = {
                                                    'tc_row': tc_row,
                                                    'tc_theme': tc_theme,
                                                    'lmrb_row': lmrb_row,
                                                    'mapped_lmrb_theme': mapped_lmrb_theme  # Store the mapped theme
                                                }
                                                
                                                st.session_state.tc_lmrb_manual_matches.append(match_entry)
                                                
                                                # Add to manual matched indices
                                                if lmrb_row not in st.session_state.manual_matched_indices:
                                                    st.session_state.manual_matched_indices.append(lmrb_row)

                                    st.success(f"Added {len(st.session_state.selected_tc_rows) * len(st.session_state.selected_lmrb_rows)} manual matches.")
                                    st.rerun()
                            else:
                                st.info("No matching LMRB entries found with current filters.")
                        else:
                            st.info("No unmatched LMRB data available.")
                    
                    with schedule_tab:
                        st.markdown("### Unmatched Schedule Data")
                        
                        if st.session_state.unmatched_schedule_data is not None and not st.session_state.unmatched_schedule_data.empty:
                            unmatched_schedule_df = st.session_state.unmatched_schedule_data
                            
                            # Schedule Theme selection for filtering
                            schedule_themes = sorted(unmatched_schedule_df['Advt_Theme'].unique())
                            
                            selected_schedule_theme = st.selectbox(
                                "Select Schedule Theme to Match",
                                options=[""] + schedule_themes,
                                key="select_schedule_theme_for_manual"
                            )
                            
                            if selected_schedule_theme:
                                st.session_state.selected_schedule_theme = selected_schedule_theme
                                
                                # Filter schedule data by selected theme
                                filtered_schedule = unmatched_schedule_df[unmatched_schedule_df['Advt_Theme'] == selected_schedule_theme].copy()
                                
                                if not filtered_schedule.empty:
                                    # Display columns
                                    display_cols = ['Advt_Theme', 'Date', 'Advt_time', 'Program', 'Dur']
                                    display_cols = [c for c in display_cols if c in filtered_schedule.columns]
                                    
                                    # Format date if needed
                                    if 'Date' not in filtered_schedule.columns and all(col in filtered_schedule.columns for col in ['Dd', 'Mn', 'Yr']):
                                        filtered_schedule['Date'] = filtered_schedule.apply(lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", axis=1)
                                    
                                    st.write(f"Unmatched schedule entries for theme '{selected_schedule_theme}':")
                                    st.dataframe(filtered_schedule[display_cols], height=200)
                                    
                                    # Find related LMRB themes
                                    related_lmrb_themes = []
                                    for mapping in st.session_state.theme_mapping:
                                        if 'schedule_theme' in mapping and mapping['schedule_theme'] == selected_schedule_theme:
                                            related_lmrb_themes.append(mapping['media_watch_theme'])
                                    
                                    # Find unmatched LMRB data with related themes
                                    if related_lmrb_themes and not unmatched_lmrb.empty:
                                        related_lmrb = unmatched_lmrb[unmatched_lmrb['Advt_Theme'].isin(related_lmrb_themes)].copy()
                                        
                                        if not related_lmrb.empty:
                                            st.markdown("### Related Unmatched LMRB Data")
                                            
                                            # Format date for display
                                            if 'Date' not in related_lmrb.columns and all(col in related_lmrb.columns for col in ['Dd', 'Mn', 'Yr']):
                                                related_lmrb['Date'] = related_lmrb.apply(lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", axis=1)
                                            
                                            display_cols = ['Advt_Theme', 'Dur', 'Date', 'Advt_time', 'Program']
                                            display_cols = [c for c in display_cols if c in related_lmrb.columns]
                                            
                                            # Display related unmatched LMRB data with row identification
                                            related_lmrb_display = related_lmrb[display_cols].copy()
                                            related_lmrb_indices = [unmatched_lmrb_indices[i] for i in range(len(related_lmrb))]
                                            related_lmrb_display.insert(0, 'LMRB_Row', related_lmrb_indices)
                                            
                                            st.dataframe(related_lmrb_display, height=200)
                                            
                                            # Create manual matching interface for schedule to LMRB
                                            st.write("Select schedule entry and LMRB entry to manually match:")
                                            
                                            col1, col2 = st.columns(2)
                                            
                                            with col1:
                                                st.write("Schedule entry:")
                                                schedule_options = []
                                                for idx, row in filtered_schedule.iterrows():
                                                    date_str = row.get('Date', f"{row['Dd']}/{row['Mn']}/{row['Yr']}" if all(col in row for col in ['Dd', 'Mn', 'Yr']) else '')
                                                    time_str = row.get('Advt_time', '')
                                                    label = f"{date_str} - {time_str}"
                                                    schedule_options.append((idx, label))
                                                
                                                selected_schedule_idx = st.selectbox(
                                                    "Select schedule entry:",
                                                    options=[x[0] for x in schedule_options],
                                                    format_func=lambda x: next((item[1] for item in schedule_options if item[0] == x), str(x)),
                                                    key="schedule_entry_selector"
                                                )
                                            
                                            with col2:
                                                st.write("LMRB entry:")
                                                lmrb_options = []
                                                for idx, row in enumerate(related_lmrb_display.itertuples()):
                                                    date_str = getattr(row, 'Date', '')
                                                    time_str = getattr(row, 'Advt_time', '')
                                                    label = f"Row {row.LMRB_Row}: {date_str} - {time_str}"
                                                    lmrb_options.append((row.LMRB_Row, label))
                                                
                                                selected_lmrb_idx = st.selectbox(
                                                    "Select LMRB entry:",
                                                    options=[x[0] for x in lmrb_options],
                                                    format_func=lambda x: next((item[1] for item in lmrb_options if item[0] == x), str(x)),
                                                    key="lmrb_entry_selector"
                                                )
                                            
                                            if st.button("Add Schedule-LMRB Match", key="add_schedule_lmrb_match"):
                                                # Check if LMRB row is already matched
                                                if selected_lmrb_idx in st.session_state.manual_matched_indices:
                                                    st.warning("This LMRB entry is already matched. Please select another one.")
                                                else:
                                                    # Add to manual matched indices
                                                    st.session_state.manual_matched_indices.append(selected_lmrb_idx)
                                                    
                                                    # Mark the schedule spot as matched
                                                    st.session_state.unmatched_schedule_data = st.session_state.unmatched_schedule_data.drop(selected_schedule_idx)
                                                    
                                                    st.success(f"Successfully matched schedule entry with LMRB row {selected_lmrb_idx}")
                                                    st.rerun()
                                        else:
                                            st.info("No related unmatched LMRB entries found.")
                                    else:
                                        st.warning(f"No LMRB themes are mapped to schedule theme '{selected_schedule_theme}'. Create a theme mapping first.")
                                else:
                                    st.info(f"No unmatched schedule entries found for theme '{selected_schedule_theme}'")
                        else:
                            st.info("No unmatched schedule data available. Run automatic matching first to identify unmatched schedule spots.")
                    
                    # Display current manual matches
                    if st.session_state.tc_lmrb_manual_matches:
                        st.markdown("---")
                        st.markdown("### Current Manual Matches")
                        
                        # Create a table of current manual matches
                        manual_matches_table = []
                        for match in st.session_state.tc_lmrb_manual_matches:
                            tc_row = match['tc_row']
                            tc_theme = match['tc_theme']
                            lmrb_row = match['lmrb_row']
                            
                            # Get TC and LMRB data
                            tc_data_filtered = tc_std[tc_std['Advt_Theme'] == tc_theme].reset_index(drop=True)
                            tc_data = tc_data_filtered.iloc[tc_row] if tc_row < len(tc_data_filtered) else None
                            
                            lmrb_data = media_watch_filtered.iloc[lmrb_row] if lmrb_row < len(media_watch_filtered) else None
                            
                            if tc_data is not None and lmrb_data is not None:
                                manual_matches_table.append({
                                    'TC Theme': tc_theme,
                                    'TC Date': f"{tc_data.get('Dd', '')}/{tc_data.get('Mn', '')}/{tc_data.get('Yr', '')}",
                                    'TC Time': tc_data.get('Advt_time', ''),
                                    'LMRB Theme': lmrb_data.get('Advt_Theme', ''),
                                    'LMRB Date': f"{lmrb_data.get('Dd', '')}/{lmrb_data.get('Mn', '')}/{lmrb_data.get('Yr', '')}",
                                    'LMRB Time': lmrb_data.get('Advt_time', ''),
                                    'TC Row': tc_row,
                                    'LMRB Row': lmrb_row
                                })
                        
                        if manual_matches_table:
                            manual_matches_df = pd.DataFrame(manual_matches_table)
                            st.dataframe(manual_matches_df, height=200)
                            
                            # Add option to remove matches
                            if st.button("Remove All Manual Matches"):
                                st.session_state.tc_lmrb_manual_matches = []
                                st.session_state.manual_matched_indices = []
                                st.success("All manual matches have been removed.")
                                st.rerun()
                            
                            # Add download button for manual matches
                            if len(st.session_state.manual_matched_indices) > 0:
                                manual_matched_data = media_watch_filtered.iloc[st.session_state.manual_matched_indices].reset_index(drop=True)
                                
                                excel_buffer = BytesIO()
                                manual_matched_data.to_excel(excel_buffer, index=False)
                                excel_buffer.seek(0)
                                
                                st.download_button(
                                    label="Download Manually Matched Data",
                                    data=excel_buffer,
                                    file_name=f"{channel}_manually_matched_data.xlsx",
                                    mime="application/vnd.ms-excel",
                                    key="download_manual_matches"
                                )
                else:
                    st.warning("Please upload both LMRB and TC data to enable manual matching.")
            
            with analytics_tab:
                with analytics_tab:
                        st.subheader("Media Reconciliation Analytics")

                        # Combine automatic and manual matches for analytics
                        combined_matches = None

                        if st.session_state.final_matched_data is not None and not st.session_state.final_matched_data.empty:
                            combined_matches = st.session_state.final_matched_data.copy()

                        # Add manual matches if they exist
                        if st.session_state.manual_matched_indices:
                            manual_matches_df = media_watch_filtered.iloc[st.session_state.manual_matched_indices].reset_index(drop=True)
                            
                            # Add match status column
                            if not manual_matches_df.empty:
                                manual_matches_df['Match_Status'] = 'Manual Match'
                                
                                # Add mapped themes if available
                                for i, idx in enumerate(st.session_state.manual_matched_indices):
                                    for match in st.session_state.tc_lmrb_manual_matches:
                                        if match.get('lmrb_row') == idx and 'mapped_lmrb_theme' in match:
                                            manual_matches_df.loc[i, 'Schedule_Theme'] = match['mapped_lmrb_theme']
                                            break
                                
                                # Combine with automatic matches
                                if combined_matches is not None:
                                    combined_matches = pd.concat([combined_matches, manual_matches_df], ignore_index=True)
                                else:
                                    combined_matches = manual_matches_df

                        if combined_matches is not None:
                            # Generate comprehensive summary
                            summary = generate_comprehensive_summary_report(
                                combined_matches, 
                                media_watch_filtered, 
                                schedule_std
                            )
                            
                            if summary is not None:
                                st.subheader("Reconciliation Summary")
                                st.dataframe(summary, height=300)
                                
                                # Create PDF summary option
                                if st.button("Generate PDF Report", key="generate_pdf_analytics"):
                                    with st.spinner("Creating PDF report... This may take a moment."):
                                        try:
                                            # Use tables-only PDF function to avoid image issues
                                            pdf_filename = create_summary_pdf_tables_only(
                                                summary,
                                                channel,
                                                combined_matches,  # Use combined matches
                                                schedule_std,
                                                st.session_state.unmatched_schedule_data
                                            )
                                            
                                            # Read the generated PDF for download
                                            with open(pdf_filename, "rb") as f:
                                                pdf_bytes = f.read()
                                            
                                            st.success(f"PDF Summary created successfully!")
                                            st.download_button(
                                                label="Download PDF Summary Report",
                                                data=pdf_bytes,
                                                file_name=pdf_filename,
                                                mime="application/pdf",
                                                key="download_pdf_analytics"
                                            )
                                            
                                            # Clean up the PDF file after download
                                            try:
                                                os.remove(pdf_filename)
                                            except:
                                                pass
                                            
                                        except Exception as e:
                                            st.error(f"Error creating PDF: {str(e)}")
                        else:
                            st.info("Run automatic matching or add manual matches to generate analytics.")
                
        else:
            st.error("Channel column not found in LMRB data.")

    else:
        st.info("Please upload LMRB data file to continue. This application helps reconcile media data across LMRB, TC, and Schedule sources.")

if __name__ == "__main__":
    main()